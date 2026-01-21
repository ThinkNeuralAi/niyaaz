"""
Sakshi.AI - Intelligent Video Analytics Platform
Main Flask Application
"""
import os
import json
import logging
from datetime import datetime, timedelta
from urllib.parse import quote_plus

# Load environment variables from .env file if available
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed, skip .env loading
from flask import Flask, render_template, request, jsonify, Response, send_from_directory, redirect, url_for, session, make_response
from io import BytesIO
from flask_sqlalchemy import SQLAlchemy
from flask_socketio import SocketIO, emit
from functools import wraps
import cv2
import threading
import time
from pathlib import Path

# Import custom modules

from modules.queue_monitor import QueueMonitor
from modules.cash_detection import CashDetection
from modules.fall_detection import FallDetection
from modules.smoking_detection import SmokingDetection
from modules.person_smoking_detection import PersonSmokingDetection
from modules.material_theft_monitor import MaterialTheftMonitor
from modules.dress_code_monitoring import DressCodeMonitoring
from modules.ppe_monitoring import PPEMonitoring
from modules.crowd_detection import CrowdDetection
from modules.table_service_monitor import TableServiceMonitor
from modules.service_discipline_monitor import ServiceDisciplineMonitor
from modules.unauthorized_entry_monitor import UnauthorizedEntryMonitor
from modules.video_processor import VideoProcessor
from modules.multi_module_processor import MultiModuleVideoProcessor
from modules.shared_multi_module_processor import SharedMultiModuleVideoProcessor
from modules.rtsp_connection_pool import rtsp_pool
from modules.database import DatabaseManager

from modules.model_manager import get_model_stats, cleanup_models

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = 'sakshi-ai-secret-key-2025'

# Database configuration
# Support both PostgreSQL (via environment variables) and SQLite (fallback)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Check for PostgreSQL environment variables first
db_host = os.getenv('DB_HOST', 'localhost')
db_port = os.getenv('DB_PORT', '5432')
db_name = os.getenv('DB_NAME', 'sakshiai')
db_user = os.getenv('DB_USER', 'postgres')
db_password = os.getenv('DB_PASSWORD', '')

# If DB_PASSWORD is set, use PostgreSQL; otherwise try config file, then fallback to SQLite
if db_password or os.getenv('USE_POSTGRESQL', '').lower() == 'true':
    # Use PostgreSQL
    if not db_password:
        # Try loading from config file as fallback
        try:
            with open(os.path.join(BASE_DIR, 'config', 'default.json'), 'r') as f:
                config = json.load(f)
                db_config = config.get('database', {})
                db_host = db_config.get('host', db_host)
                db_port = str(db_config.get('port', db_port))
                db_name = db_config.get('name', db_name)
                db_user = db_config.get('username', db_user)
                db_password = db_config.get('password', '')
        except Exception as e:
            logger.warning(f"Could not load database config from file: {e}")
    
    if db_password:
        # Construct PostgreSQL connection string
        # URL-encode password to handle special characters like @, :, etc.
        encoded_password = quote_plus(db_password)
        app.config['SQLALCHEMY_DATABASE_URI'] = f"postgresql://{db_user}:{encoded_password}@{db_host}:{db_port}/{db_name}"
        logger.info(f"Using PostgreSQL database: {db_name}@{db_host}:{db_port}")
    else:
        # Fallback to SQLite if no password provided
        db_path = os.path.join(BASE_DIR, "data", "sakshi.db")
        app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{db_path}"
        logger.info(f"Using SQLite database: {db_path}")
else:
    # Use SQLite as default
    db_path = os.path.join(BASE_DIR, "data", "sakshi.db")
    app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{db_path}"
    logger.info(f"Using SQLite database: {db_path}")

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize extensions
db = SQLAlchemy(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Initialize Telegram notifier at startup to verify configuration
try:
    from modules.telegram_notifier import get_telegram_notifier
    telegram_notifier = get_telegram_notifier()
    if telegram_notifier.enabled:
        logger.info("✅ Telegram notifications ENABLED and ready")
    else:
        logger.warning("⚠️ Telegram notifications DISABLED - check environment variables TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID")
except Exception as e:
    logger.warning(f"⚠️ Could not initialize Telegram notifier: {e}")

# Global variables - Updated for multi-module support
shared_video_processors = {}  # {channel_id: MultiModuleVideoProcessor}
channel_modules = {}  # {channel_id: {module_name: module_instance}}
app_configs = {
    # 1) Queue & wait time tracking
    'QueueMonitor': {
        'name': 'Queue & Wait Time',
        'description': 'Monitor queue length, wait time, and counter staffing',
        'channels': {},
        'status': 'online'
    },
    # 2) Uniform compliance
    'DressCodeMonitoring': {
        'name': 'Uniform Compliance',
        'description': 'Monitor employee uniform compliance',
        'channels': {},
        'status': 'online'
    },
    # 2b) PPE compliance (separate from uniform)
    'PPEMonitoring': {
        'name': 'PPE Compliance',
        'description': 'Monitor Personal Protective Equipment compliance (Apron, Gloves, Hairnet)',
        'channels': {},
        'status': 'online'
    },
    # 3) Cash drawer monitoring
    'CashDetection': {
        'name': 'Cash Drawer Monitoring',
        'description': 'Detect cash and open drawers to monitor transactions',
        'channels': {},
        'status': 'online'
    },
    # 4) Smoke & fire detection (reusing smoking/safety model later if needed)
    'SmokingDetection': {
        'name': 'Smoke & Fire Detection',
        'description': 'Detect smoke and fire events for safety alerts',
        'channels': {},
        'status': 'online'
    },
    # 5) Crowd detection in parking space
    'CrowdDetection': {
        'name': 'Crowd Detection',
        'description': 'Monitor crowd gathering in parking space',
        'channels': {},
        'status': 'online'
    },
    # 6) Table service monitoring
    'TableServiceMonitor': {
        'name': 'Table Cleanliness',
        'description': 'Monitor table cleanliness and reset times after customers leave',
        'channels': {},
        'status': 'online'
    },
    'ServiceDisciplineMonitor': {
        'name': 'Service Discipline',
        'description': 'Order wait time: seated to first waiter visit',
        'channels': {},
        'status': 'online'
    },
    # 7) Fall detection
    'FallDetection': {
        'name': 'Fall Detection',
        'description': 'Detect person falls using advanced tracking and history-based analysis',
        'channels': {},
        'status': 'online'
    },
    # 8) Material Theft / Misuse
    'MaterialTheftMonitor': {
        'name': 'Material Theft / Misuse',
        'description': 'Monitor for material theft/misuse on weighing machine',
        'channels': {},
        'status': 'online'
    },
}

# Database manager
db_manager = DatabaseManager(db)

# ============= Channel Auto-Loader from Configuration =============
def load_channels_from_config(config_file='config/channels.json'):
    """
    Load and start channels from config file on application startup
    
    Args:
        config_file: Path to the channels configuration file (default: config/channels.json)
    """
    config_path = Path(config_file)
    
    if not config_path.exists():
        logger.warning(f"Channel configuration file not found: {config_path}")
        return
    
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
        
        channels = config.get('channels', [])
        logger.info(f"Loading {len(channels)} channels from configuration...")
        
        for channel_config in channels:
            try:
                if not channel_config.get('enabled', False):
                    logger.info(f"Skipping disabled channel: {channel_config.get('channel_name', 'Unknown')}")
                    continue
                
                channel_id = channel_config['channel_id']
                channel_name = channel_config['channel_name']
                # Support both RTSP URLs and local video files
                rtsp_url = channel_config.get('rtsp_url', '')
                video_file = channel_config.get('video_file', '')
                video_source = video_file if video_file else rtsp_url
                modules_config = channel_config.get('modules', [])
                
                # Save/update channel to database for persistence
                try:
                    with app.app_context():
                        existing = db_manager.get_rtsp_channel(channel_id)
                        # Always update RTSP URL from config file to ensure it's current
                        db_manager.save_rtsp_channel(channel_id, channel_name, video_source, 
                                                   description=f"Auto-loaded from channels.json")
                        if existing:
                            logger.info(f"  💾 Updated channel '{channel_name}' RTSP URL in database")
                        else:
                            logger.info(f"  💾 Saved channel '{channel_name}' to database")
                except Exception as e:
                    logger.warning(f"  ⚠ Could not save channel to database: {e}")
                
                # Determine video source with fallback logic
                # Try RTSP first, fall back to video_file if RTSP fails
                final_video_source = None
                source_type = "unknown"
                
                if rtsp_url:
                    # Try RTSP first
                    logger.info(f"Starting channel '{channel_name}' ({channel_id}) with {len(modules_config)} modules (trying RTSP first)")
                    final_video_source = rtsp_url
                    source_type = "RTSP"
                elif video_file:
                    # Use video file directly if no RTSP URL
                    logger.info(f"Starting channel '{channel_name}' ({channel_id}) with {len(modules_config)} modules (using video file)")
                    final_video_source = video_file
                    source_type = "video file"
                else:
                    logger.warning(f"Channel '{channel_name}' ({channel_id}) has no RTSP URL or video file - skipping")
                    continue
                
                # Create video processor if it doesn't exist
                if channel_id not in shared_video_processors:
                    # Use SharedMultiModuleVideoProcessor - supports both RTSP and local video files
                    processor = SharedMultiModuleVideoProcessor(
                        video_source=final_video_source,
                        channel_id=channel_id,
                        fps_limit=30
                    )
                    shared_video_processors[channel_id] = processor
                    
                    # Initialize channel modules dictionary
                    if channel_id not in channel_modules:
                        channel_modules[channel_id] = {}
                
                # Add each module to the processor
                for module_config in modules_config:
                    module_type = module_config['type']
                    module_settings = module_config.get('config', {})
                    
                    try:
                        # Create module instance based on type
                        if module_type == 'PeopleCounter':
                            module = PeopleCounter(channel_id, socketio, db_manager, app)
                            # Apply counting line configuration if provided
                            if 'counting_line' in module_settings:
                                logger.info(f"  📏 Loading counting line for {channel_id}:")
                                logger.info(f"     {json.dumps(module_settings['counting_line'], indent=6)}")
                                module.set_counting_line(module_settings['counting_line'])
                        
                        elif module_type == 'QueueMonitor':
                            module = QueueMonitor(channel_id, socketio, db_manager, app)
                            # Load configuration from database first
                            try:
                                with app.app_context():
                                    module.load_configuration()
                            except Exception as e:
                                logger.warning(f"Could not load QueueMonitor config from DB: {e}")
                            # Apply ROI configuration from config file if provided (overrides DB)
                            if module_settings:
                                # Map the channels.json structure to QueueMonitor format
                                roi_points = {
                                    'main': module_settings.get('queue_roi', {}).get('points', []),
                                    'secondary': module_settings.get('counter_roi', {}).get('points', [])
                                }
                                if roi_points['main'] or roi_points['secondary']:
                                    logger.info(f"  📐 Loading ROI configuration from config file for {channel_id}:")
                                    logger.info(f"     Queue area: {len(roi_points['main'])} points")
                                    logger.info(f"     Counter area: {len(roi_points['secondary'])} points")
                                    module.set_roi(roi_points)
                            # Apply settings if provided
                            if 'settings' in module_settings:
                                module.settings.update(module_settings['settings'])
                        
                        elif module_type == 'BagDetection':
                            module = BagDetection(channel_id, socketio, db_manager, app)
                            # Apply settings if provided
                            if module_settings:
                                for key, value in module_settings.items():
                                    if hasattr(module, key):
                                        setattr(module, key, value)
                        
                        elif module_type == 'CashDetection':
                            module = CashDetection(channel_id, socketio, db_manager, app)
                            # Apply settings if provided
                            if module_settings:
                                for key, value in module_settings.items():
                                    if hasattr(module, key):
                                        setattr(module, key, value)
                        
                        elif module_type == 'FallDetection':
                            module = FallDetection(channel_id, socketio, db_manager, app)
                            # Apply settings if provided
                            if module_settings:
                                for key, value in module_settings.items():
                                    if hasattr(module, key):
                                        setattr(module, key, value)
                        
                        elif module_type == 'MoppingDetection':
                            module = MoppingDetection(channel_id, socketio, db_manager, app)
                            # Apply settings if provided
                            if module_settings:
                                for key, value in module_settings.items():
                                    if hasattr(module, key):
                                        setattr(module, key, value)
                        
                        elif module_type == 'SmokingDetection':
                            module = SmokingDetection(channel_id, socketio, db_manager, app)
                            # Apply settings if provided
                            if module_settings:
                                for key, value in module_settings.items():
                                    if hasattr(module, key):
                                        setattr(module, key, value)
                        
                        elif module_type == 'PersonSmokingDetection':
                            module = PersonSmokingDetection(channel_id, socketio, db_manager, app)
                            logger.info(f"  ✓ Added Smoking Detection to channel {channel_id}")
                            # Apply settings if provided
                            if module_settings:
                                if 'alert_cooldown' in module_settings:
                                    module.alert_cooldown = float(module_settings['alert_cooldown'])
                                if 'conf_threshold' in module_settings:
                                    module.conf_threshold = float(module_settings['conf_threshold'])
                                if 'detection_duration_threshold' in module_settings:
                                    module.detection_duration_threshold = float(module_settings['detection_duration_threshold'])
                                logger.info(f"  ✅ Loaded settings from config for {channel_id}")
                        
                        elif module_type == 'PhoneUsageDetection':
                            module = PhoneUsageDetection(channel_id, socketio, db_manager, app)
                            # Apply settings if provided
                            if module_settings:
                                for key, value in module_settings.items():
                                    if hasattr(module, key):
                                        setattr(module, key, value)
                        
                        elif module_type == 'RestrictedAreaMonitor':
                            model_path = module_settings.get('model_path', 'models/best.pt') if module_settings else 'models/best.pt'
                            module = RestrictedAreaMonitor(channel_id, model_path, db_manager, socketio, module_settings)
                            
                            # Load ROI points - try config first, then database
                            roi_loaded = False
                            if module_settings and 'roi_points' in module_settings:
                                module.set_roi_points(module_settings['roi_points'])
                                roi_loaded = True
                                logger.info(f"✅ Loaded ROI from config file for {channel_id}")
                            
                            # If not in config, try loading from database (with app context)
                            if not roi_loaded:
                                try:
                                    with app.app_context():
                                        saved_roi = db_manager.get_channel_config(channel_id, 'RestrictedAreaMonitor', 'roi')
                                        if saved_roi:
                                            # Handle both dict and list formats
                                            if isinstance(saved_roi, dict):
                                                if 'main' in saved_roi:
                                                    module.set_roi_points(saved_roi['main'])
                                                    logger.info(f"✅ Loaded ROI from database (dict format) for {channel_id}: {len(saved_roi['main'])} points")
                                                else:
                                                    logger.warning(f"⚠️ ROI dict has no 'main' key for {channel_id}: {saved_roi}")
                                            elif isinstance(saved_roi, list):
                                                module.set_roi_points(saved_roi)
                                                logger.info(f"✅ Loaded ROI from database (list format) for {channel_id}: {len(saved_roi)} points")
                                            else:
                                                logger.warning(f"⚠️ Unexpected ROI format for {channel_id}: {type(saved_roi)}")
                                        else:
                                            logger.info(f"ℹ️ No saved ROI found in database for {channel_id}")
                                except Exception as e:
                                    logger.error(f"❌ Could not load ROI from database for {channel_id}: {e}", exc_info=True)
                        
                        elif module_type == 'DressCodeMonitoring':
                            module = DressCodeMonitoring(channel_id, socketio, db_manager, app)
                            
                            # Load counter ROI from config file (if provided)
                            if module_settings:
                                counter_roi_points = module_settings.get('counter_roi', {})
                                if isinstance(counter_roi_points, dict) and 'points' in counter_roi_points:
                                    module.set_counter_roi(counter_roi_points['points'])
                                    logger.info(f"  📐 Loaded counter ROI from config for {channel_id}: {len(counter_roi_points['points'])} points")
                                elif isinstance(counter_roi_points, list):
                                    module.set_counter_roi(counter_roi_points)
                                    logger.info(f"  📐 Loaded counter ROI from config for {channel_id}: {len(counter_roi_points)} points")
                                
                                # Load allowed uniforms from config
                                allowed_uniforms = module_settings.get('allowed_uniforms', {})
                                if allowed_uniforms:
                                    module.set_allowed_uniforms(allowed_uniforms)
                                    logger.info(f"  📐 Loaded allowed uniforms from config for {channel_id}: {allowed_uniforms}")
                                
                                # Apply other settings
                                for key, value in module_settings.items():
                                    if key not in ['counter_roi', 'allowed_uniforms'] and hasattr(module, key):
                                        setattr(module, key, value)
                        
                        elif module_type == 'TableServiceMonitor':
                            module = TableServiceMonitor(channel_id, socketio, db_manager, app)
                            
                            # Load table ROIs from config file (if provided)
                            if module_settings:
                                table_rois = module_settings.get('table_rois', {})
                                if table_rois:
                                    for table_id, roi_data in table_rois.items():
                                        if isinstance(roi_data, dict) and 'points' in roi_data:
                                            module.set_table_roi(table_id, roi_data['points'])
                                        elif isinstance(roi_data, list):
                                            module.set_table_roi(table_id, roi_data)
                                    logger.info(f"  📐 Loaded {len(table_rois)} table ROIs from config for {channel_id}")
                                
                                # Load settings from config
                                if 'settings' in module_settings:
                                    module.settings.update(module_settings['settings'])
                                    logger.info(f"  ✅ Loaded settings from config for {channel_id}")

                        elif module_type == 'ServiceDisciplineMonitor':
                            module = ServiceDisciplineMonitor(channel_id, socketio, db_manager, app)
                            logger.info(f"  ✓ Added ServiceDisciplineMonitor to channel {channel_id}")
                            # Load from DB
                            try:
                                with app.app_context():
                                    module.load_configuration()
                            except Exception as e:
                                logger.warning(f"Could not load ServiceDisciplineMonitor config from DB: {e}")
                            # Apply ROIs from config
                            if module_settings:
                                table_rois = module_settings.get('table_rois', {})
                                if isinstance(table_rois, dict):
                                    for table_id, roi_data in table_rois.items():
                                        if isinstance(roi_data, dict) and 'points' in roi_data:
                                            module.set_table_roi(table_id, roi_data['points'])
                                        elif isinstance(roi_data, list):
                                            module.set_table_roi(table_id, roi_data)
                                    logger.info(f"  📐 Loaded {len(table_rois)} table ROIs from config for {channel_id}")
                                if 'settings' in module_settings:
                                    module.settings.update(module_settings['settings'])
                                    logger.info(f"  ✅ Loaded settings from config for {channel_id}")
                        
                        elif module_type == 'UnauthorizedEntryMonitor':
                            module = UnauthorizedEntryMonitor(channel_id, socketio, db_manager, app)
                            logger.info(f"  ✓ Added UnauthorizedEntryMonitor to channel {channel_id}")
                            # Apply settings if provided
                            if module_settings:
                                if 'alert_cooldown' in module_settings:
                                    module.alert_cooldown = float(module_settings['alert_cooldown'])
                                if 'conf_threshold' in module_settings:
                                    module.conf_threshold = float(module_settings['conf_threshold'])
                                logger.info(f"  ✅ Loaded settings from config for {channel_id}")

                        elif module_type == 'MaterialTheftMonitor':
                            module = MaterialTheftMonitor(channel_id, socketio, db_manager, app, module_settings or {})
                            logger.info(f"  ✓ Added MaterialTheftMonitor to channel {channel_id}")
                            # Load ROI configuration from config file if provided
                            if module_settings:
                                roi_points = module_settings.get('roi_points', [])
                                if roi_points:
                                    # Check if points are normalized (0-1) or absolute pixels
                                    if roi_points and len(roi_points) > 0:
                                        # If first point values are > 1, they're absolute pixels (need normalization)
                                        # Otherwise assume normalized
                                        first_point = roi_points[0]
                                        if isinstance(first_point, list) and len(first_point) >= 2:
                                            if first_point[0] > 1 or first_point[1] > 1:
                                                logger.warning(f"  ⚠️ MaterialTheftMonitor ROI points appear to be absolute pixels. They will be normalized on first frame.")
                                    module.set_roi(roi_points)
                                    logger.info(f"  📐 Loaded ROI from config for {channel_id}: {len(roi_points)} points")
                                
                                # Load other settings
                                if 'min_area' in module_settings:
                                    module.min_area = int(module_settings['min_area'])
                                if 'still_frames_required' in module_settings:
                                    module.still_frames_required = int(module_settings['still_frames_required'])
                                if 'alert_cooldown' in module_settings:
                                    module.alert_cooldown = float(module_settings['alert_cooldown'])
                                logger.info(f"  ✅ Loaded MaterialTheftMonitor settings from config for {channel_id}")
                        
                        elif module_type == 'PPEMonitoring':
                            module = PPEMonitoring(channel_id, socketio, db_manager, app)
                            logger.info(f" ✓ Added PPEMonitoring to channel {channel_id}")
                            # Apply settings if provided
                            if module_settings:
                                if 'required_items' in module_settings:
                                    module.set_settings({'required_items': module_settings['required_items']})
                                    logger.info(f"   Applied required_items: {module_settings['required_items']}")
                                if 'settings' in module_settings:
                                    module.set_settings(module_settings['settings'])
                                    logger.info(f"   Applied settings: {module_settings['settings']}")
                                # Also apply individual settings
                                for key, value in module_settings.items():
                                    if key not in ['required_items', 'settings'] and hasattr(module, key):
                                        setattr(module, key, value)
                        
                        elif module_type == 'CrowdDetection':
                            module = CrowdDetection(channel_id, socketio, db_manager, app)
                            # Load configuration from database first
                            try:
                                with app.app_context():
                                    module.load_configuration()
                            except Exception as e:
                                logger.warning(f"Could not load CrowdDetection config from DB: {e}")
                            # Apply ROI configuration from config file if provided (overrides DB)
                            if module_settings:
                                roi_points = module_settings.get('roi', {})
                                if isinstance(roi_points, dict) and 'points' in roi_points:
                                    module.set_roi({'main': roi_points['points']})
                                elif isinstance(roi_points, list):
                                    module.set_roi({'main': roi_points})
                            # Apply settings if provided
                            if 'settings' in module_settings:
                                module.set_settings(module_settings['settings'])
                        
                        elif module_type == 'TableServiceMonitor':
                            module = TableServiceMonitor(channel_id, socketio, db_manager, app)
                            logger.info(f"  ✓ Added TableServiceMonitor to channel {channel_id}")
                            # Load configuration from database first
                            try:
                                with app.app_context():
                                    module.load_configuration()
                            except Exception as e:
                                logger.warning(f"Could not load TableServiceMonitor config from DB: {e}")
                            # Apply table ROIs from config file if provided
                            if module_settings:
                                table_rois = module_settings.get('table_rois', {})
                                if isinstance(table_rois, dict):
                                    for table_id, roi_data in table_rois.items():
                                        if isinstance(roi_data, dict) and 'points' in roi_data:
                                            module.set_table_roi(table_id, roi_data['points'])
                                        elif isinstance(roi_data, list):
                                            module.set_table_roi(table_id, roi_data)
                                    logger.info(f"  📐 Loaded {len(table_rois)} table ROIs from config for {channel_id}")
                                
                                # Load settings from config
                                if 'settings' in module_settings:
                                    module.settings.update(module_settings['settings'])
                                    logger.info(f"  ✅ Loaded settings from config for {channel_id}")
                        
                        elif module_type == 'Heatmap':
                            module = HeatmapProcessor(channel_id, socketio, db_manager, app)
                            # Apply settings if provided
                            if module_settings:
                                for key, value in module_settings.items():
                                    if hasattr(module, key):
                                        setattr(module, key, value)
                        
                        else:
                            logger.warning(f"Unknown module type: {module_type}")
                            continue
                        
                        # Add module to processor and track it
                        logger.info(f"  🔧 Adding {module_type} to processor for {channel_id}")
                        shared_video_processors[channel_id].add_module(module_type, module)
                        channel_modules[channel_id][module_type] = module
                        logger.info(f"  ✅ {module_type} added. Active modules now: {shared_video_processors[channel_id].get_active_modules()}")
                        
                        # Also populate app_configs so modules loaded from
                        # channels.json appear correctly in the dashboard for
                        # their respective apps (not only when started via
                        # /api/start_channel).
                        if module_type in app_configs:
                            if channel_id not in app_configs[module_type]['channels']:
                                app_configs[module_type]['channels'][channel_id] = {
                                    'name': channel_name,
                                    'status': 'online',
                                    'video_source': rtsp_url,
                                    'source_type': 'rtsp',
                                    'shared': True,
                                    'active_modules': []
                                }
                            active = app_configs[module_type]['channels'][channel_id].setdefault('active_modules', [])
                            if module_type not in active:
                                active.append(module_type)
                        
                        logger.info(f"  ✓ Added {module_type} to channel {channel_id}")
                        
                    except Exception as e:
                        logger.error(f"Failed to add module {module_type} to channel {channel_id}: {e}")
                        continue
                
                # Start the video processor
                try:
                    start_result = shared_video_processors[channel_id].start()
                    logger.info(f"Channel '{channel_name}' start() returned: {start_result}")
                    if start_result:
                        logger.info(f"✓ Channel '{channel_name}' started successfully")
                    else:
                        # RTSP connection failed - try fallback to video file if available
                        if rtsp_url and video_file:
                            logger.warning(f"⚠ RTSP connection failed for '{channel_name}', trying fallback to video file: {video_file}")
                            
                            # Remove failed processor
                            if channel_id in shared_video_processors:
                                failed_processor = shared_video_processors[channel_id]
                                try:
                                    failed_processor.stop()
                                except:
                                    pass
                                del shared_video_processors[channel_id]
                            
                            # Create new processor with video file
                            try:
                                fallback_processor = SharedMultiModuleVideoProcessor(
                                    video_source=video_file,
                                    channel_id=channel_id,
                                    fps_limit=30
                                )
                                shared_video_processors[channel_id] = fallback_processor
                                
                                # Re-add all modules to the new processor
                                for module_name, module_instance in channel_modules[channel_id].items():
                                    fallback_processor.add_module(module_name, module_instance)
                                
                                # Try to start with video file
                                fallback_result = fallback_processor.start()
                                if fallback_result:
                                    logger.info(f"✅ Channel '{channel_name}' started successfully using video file fallback")
                                else:
                                    logger.warning(f"⚠ Video file fallback also failed for '{channel_name}'")
                                    logger.warning(f"⚠ Keeping channel in channel_modules for dashboard visibility (is_running=False)")
                                    if channel_id in shared_video_processors:
                                        del shared_video_processors[channel_id]
                            except Exception as fallback_error:
                                logger.error(f"❌ Video file fallback failed for '{channel_name}': {fallback_error}")
                                logger.warning(f"⚠ Keeping channel in channel_modules for dashboard visibility (is_running=False)")
                                if channel_id in shared_video_processors:
                                    del shared_video_processors[channel_id]
                        else:
                            # start() may return False if the processor is already running.
                            # Treat that case as success and do NOT delete the processor.
                            processor = shared_video_processors.get(channel_id)
                            thread_alive = bool(getattr(processor, 'processing_thread', None) and processor.processing_thread.is_alive()) if processor else False
                            has_frame = bool(getattr(processor, 'latest_raw_frame', None) is not None or getattr(processor, 'latest_annotated_frame', None) is not None) if processor else False
                            if processor and (getattr(processor, 'is_running', False) or thread_alive or has_frame):
                                logger.info(f"✓ Channel '{channel_name}' is already running (thread_alive={thread_alive}, has_frame={has_frame}); keeping existing processor")
                            else:
                                logger.warning(f"⚠ Channel '{channel_name}' processor start() returned False - RTSP connection may have failed")
                                logger.warning(f"⚠ No video file fallback available (video_file not configured)")
                                logger.warning(f"⚠ Keeping channel in channel_modules for dashboard visibility (is_running=False)")
                                # Remove processor from shared_video_processors since it's not actually running
                                if channel_id in shared_video_processors:
                                    del shared_video_processors[channel_id]
                                    logger.info(f"Removed non-running processor for {channel_id} from shared_video_processors")
                except Exception as e:
                    logger.error(f"❌ Failed to start channel '{channel_name}': {e}", exc_info=True)
                    # Keep channel_modules entry even on failure so dashboard can show it
                    # Only remove processor if it exists
                    if channel_id in shared_video_processors:
                        del shared_video_processors[channel_id]
                    # Keep channel_modules so dashboard can show configured channels
                
            except KeyError as e:
                logger.error(f"Invalid channel configuration - missing required field: {e}")
                continue
            except Exception as e:
                logger.error(f"Error loading channel: {e}")
                continue
        
        # Summary of channel loading
        running_count = len(shared_video_processors)
        configured_count = len(channel_modules)
        logger.info(f"📊 Channel loading complete:")
        logger.info(f"   - {running_count} channels with running processors")
        logger.info(f"   - {configured_count} channels configured (may include non-running)")
        
        # Log which channels are running vs configured
        running_channels = set(shared_video_processors.keys())
        configured_channels = set(channel_modules.keys())
        not_running = configured_channels - running_channels
        if not_running:
            logger.warning(f"   ⚠ {len(not_running)} channels configured but not running: {sorted(not_running)}")
        if running_channels:
            logger.info(f"   ✓ {len(running_channels)} channels running: {sorted(running_channels)}")
        
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse channel configuration file: {e}")
    except Exception as e:
        logger.error(f"Error loading channels from config: {e}")

# ============= Authentication Decorator =============
def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            # Check if this is an API request (JSON expected)
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'error': 'Authentication required'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Note: Removed admin_required decorator - all users now have access

# ============= Routes =============

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page (DEV MODE: auto-login)"""
    if request.method == 'POST':
        username = request.form.get('username') or 'devuser'

        # 🚨 Dev-only: skip DB check and just log in
        session['user_id'] = 1
        session['username'] = username
        session['role'] = 'admin'

        return redirect(url_for('dashboard'))

    # If already logged in, redirect to dashboard
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    return render_template('login.html')


@app.route('/logout')
def logout():
    """Logout user"""
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def landing():
    """Landing page"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard with full analytics (original UI)"""
    user = db_manager.get_user_by_id(session['user_id'])
    return render_template('dashboard.html', app_configs=app_configs, user=user)


@app.route('/dashboard_clean')
@login_required
def dashboard_clean():
    """Minimal dashboard view (clean 4-module UI)"""
    user = db_manager.get_user_by_id(session['user_id'])
    return render_template('clean_dashboard.html', app_configs=app_configs, user=user)

@app.route('/static/alerts/<filename>')
@login_required
def serve_alert_gif(filename):
    """Serve alert GIF files"""
    return send_from_directory('static/alerts', filename)

@app.route('/api/get_current_user')
@login_required
def get_current_user():
    """Get current logged in user info"""
    return jsonify({
        'success': True,
        'user': {
            'id': session.get('user_id'),
            'username': session.get('username'),
            'role': session.get('role')
        }
    })

@app.route('/video_feed/<app_name>/<channel_id>')
@login_required
def video_feed(app_name, channel_id):
    """Video streaming endpoint - supports both shared and module-specific feeds"""
    def generate():
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            while processor.is_running:
                # Get frame - either combined or module-specific
                frame = processor.get_latest_frame(module_name=app_name if app_name in processor.get_active_modules() else None)
                if frame is not None:
                    # Encode with optimized JPEG quality for faster transmission
                    ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 75])  # 75% quality (vs 95% default)
                    if ret:
                        yield (b'--frame\r\n'
                               b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
                time.sleep(0.066)  # ~15 FPS (reduced from 10 FPS for smoother streaming)
    
    return Response(generate(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/get_channels/<app_name>')
def get_channels(app_name):
    """Get available channels for an app"""
    channels = []
    
    # Get saved RTSP channels from database
    try:
        saved_channels = db_manager.get_rtsp_channels()
        for channel in saved_channels:
            channels.append({
                'id': channel['channel_id'],
                'name': channel['name'],
                'rtsp_url': channel['rtsp_url'],
                'type': 'rtsp'
            })
    except Exception as e:
        logger.error(f"Error loading RTSP channels: {e}")
    
    # Also check for video files (backward compatibility)
    videos_dir = Path('videos')
    if videos_dir.exists():
        for video_file in videos_dir.glob('*.mp4'):
            channel_id = video_file.stem
            channels.append({
                'id': channel_id,
                'name': f"Video {channel_id}",
                'path': str(video_file),
                'type': 'video'
            })
    
    return jsonify(channels)

@app.route('/api/get_active_channels')
def get_active_channels():
    """Get all currently active channels with their running modules"""
    try:
        active_channels = []
        
        logger.info(f"📊 get_active_channels: shared_video_processors has {len(shared_video_processors)} entries")
        logger.info(f"📊 get_active_channels: channel_modules has {len(channel_modules)} entries")
        
        # Only get channels from processors that are actually running and can provide frames
        for channel_id, processor in shared_video_processors.items():
            # Check if processor is actually running
            is_running = getattr(processor, 'is_running', False)
            
            # Check if processor can actually get frames (has valid connection)
            has_frames = False
            try:
                if hasattr(processor, 'get_latest_frame'):
                    test_frame = processor.get_latest_frame()
                    # Check if frame is valid (not None and not empty)
                    if test_frame is not None:
                        if hasattr(test_frame, 'size'):
                            has_frames = test_frame.size > 0
                        else:
                            has_frames = True  # Non-array frame, assume valid
            except Exception as e:
                logger.debug(f"Channel {channel_id}: Could not get frame: {e}")
                has_frames = False
            
            # Only include if actually running AND can provide frames
            if is_running and has_frames:
                # Check if channel has modules configured
                if channel_id in channel_modules and channel_modules[channel_id]:
                    # Get active modules for this channel
                    active_modules = processor.get_active_modules() if hasattr(processor, 'get_active_modules') else list(channel_modules[channel_id].keys())
                    
                    logger.info(f"📊 Channel {channel_id}: ACTIVE - modules={active_modules}, is_running={is_running}, has_frames={has_frames}")
                    
                    active_channels.append({
                        'channel_id': channel_id,
                        'modules': active_modules,
                        'is_running': True
                    })
                else:
                    logger.debug(f"📊 Channel {channel_id}: Running but no modules configured")
            else:
                logger.debug(f"📊 Channel {channel_id}: NOT ACTIVE - is_running={is_running}, has_frames={has_frames}")
        
        # Don't include channels that are just configured but not running
        # (They will appear when the processor actually starts)
        
        logger.info(f"📊 Returning {len(active_channels)} ACTIVE channels (only running with valid frames)")
        
        return jsonify({
            'success': True,
            'active_channels': active_channels,
            'count': len(active_channels)
        })
    except Exception as e:
        logger.error(f"Error getting active channels: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_configured_channels')
def get_configured_channels():
    """Get all configured channels from channels.json (for fallback when active channels aren't loaded yet)"""
    try:
        config_path = Path('config/channels.json')
        if not config_path.exists():
            return jsonify({'success': False, 'error': 'channels.json not found'})
        
        with open(config_path, 'r') as f:
            config = json.load(f)
        
        channels = config.get('channels', [])
        # Return channels with their module types
        configured_channels = []
        for ch in channels:
            if ch.get('enabled', False):
                modules = ch.get('modules', [])
                module_types = [m.get('type') for m in modules if isinstance(m, dict) and m.get('type')]
                configured_channels.append({
                    'channel_id': ch.get('channel_id'),
                    'channel_name': ch.get('channel_name'),
                    'modules': module_types,
                    'enabled': ch.get('enabled', True)
                })
        
        return jsonify({
            'success': True,
            'channels': configured_channels,
            'count': len(configured_channels)
        })
    except Exception as e:
        logger.error(f"Error getting configured channels: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/add_rtsp_channel', methods=['POST'])
@login_required
def add_rtsp_channel():
    """Add a new RTSP channel"""
    data = request.json
    channel_name = data.get('name')
    rtsp_url = data.get('rtsp_url')
    
    if not channel_name or not rtsp_url:
        return jsonify({'success': False, 'error': 'Name and RTSP URL are required'})
    
    try:
        # Generate channel ID from name
        channel_id = channel_name.lower().replace(' ', '_').replace('-', '_')
        
        # Save to database
        success = db_manager.save_rtsp_channel(channel_id, channel_name, rtsp_url)
        
        if success:
            return jsonify({
                'success': True, 
                'channel_id': channel_id,
                'message': f'RTSP channel "{channel_name}" added successfully'
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to save RTSP channel'})
    
    except Exception as e:
        logger.error(f"Error adding RTSP channel: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/remove_rtsp_channel', methods=['POST'])
@login_required
def remove_rtsp_channel():
    """Remove an RTSP channel"""
    data = request.json
    channel_id = data.get('channel_id')
    
    if not channel_id:
        return jsonify({'success': False, 'error': 'Channel ID is required'})
    
    try:
        # Stop channel if running
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            processor.stop()
            del shared_video_processors[channel_id]
            if channel_id in channel_modules:
                del channel_modules[channel_id]
        
        # Remove from database
        success = db_manager.remove_rtsp_channel(channel_id)
        
        if success:
            return jsonify({'success': True, 'message': 'RTSP channel removed successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to remove RTSP channel'})
    
    except Exception as e:
        logger.error(f"Error removing RTSP channel: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/start_channel', methods=['POST'])
@login_required
def start_channel():
    """Start processing a video channel - supports multiple modules on same video"""
    data = request.json
    app_name = data.get('app_name')
    channel_id = data.get('channel_id')
    
    # Determine video source (RTSP URL or video file path)
    rtsp_url = data.get('rtsp_url')
    video_path = data.get('video_path')
    
    if rtsp_url:
        video_source = rtsp_url
        source_type = 'rtsp'
    elif video_path:
        video_source = video_path
        source_type = 'video'
    else:
        # Try to get from database or fallback to video file
        try:
            rtsp_channel = db_manager.get_rtsp_channel(channel_id)
            if rtsp_channel:
                video_source = rtsp_channel['rtsp_url']
                source_type = 'rtsp'
            else:
                video_source = f'videos/{channel_id}.mp4'
                source_type = 'video'
        except:
            video_source = f'videos/{channel_id}.mp4'
            source_type = 'video'
    
    try:
        # Check if video processor already exists for this channel
        if channel_id not in shared_video_processors:
            # Create new multi-module processor
            processor = MultiModuleVideoProcessor(video_source, channel_id)
            shared_video_processors[channel_id] = processor
            channel_modules[channel_id] = {}
            
            # Start the processor
            if not processor.start():
                del shared_video_processors[channel_id]
                del channel_modules[channel_id]
                return jsonify({'success': False, 'error': f'Failed to start video processor for {source_type} source'})
        
        processor = shared_video_processors[channel_id]
        
        # Check if this module is already active for this channel
        if app_name in channel_modules[channel_id]:
            return jsonify({'success': True, 'message': f'{app_name} already active on channel {channel_id}'})
        
        # Create and add the analysis module
        if app_name == 'PeopleCounter':
            module = PeopleCounter(channel_id, socketio, db_manager, app)
        elif app_name == 'QueueMonitor':
            module = QueueMonitor(channel_id, socketio, db_manager, app)
            # Load saved ROI configuration from database
            try:
                with app.app_context():
                    module.load_configuration()
                logger.info(f"✅ Loaded QueueMonitor configuration from database for {channel_id}")
            except Exception as e:
                logger.warning(f"⚠️ Could not load QueueMonitor config from DB for {channel_id}: {e}")
        elif app_name == 'BagDetection':
            module = BagDetection(channel_id, socketio, db_manager, app)
        elif app_name == 'Heatmap':
            module = HeatmapProcessor(channel_id, socketio, db_manager, app)
        elif app_name == 'CashDetection':
            module = CashDetection(channel_id, socketio, db_manager, app)
        elif app_name == 'FallDetection':
            module = FallDetection(channel_id, socketio, db_manager, app)
        elif app_name == 'MoppingDetection':
            module = MoppingDetection(channel_id, socketio, db_manager, app)
        elif app_name == 'SmokingDetection':
            module = SmokingDetection(channel_id, socketio, db_manager, app)
        elif app_name == 'PhoneUsageDetection':
            module = PhoneUsageDetection(channel_id, socketio, db_manager, app)
        elif app_name == 'RestrictedAreaMonitor':
            model_path = 'models/best.pt'
            module = RestrictedAreaMonitor(channel_id, model_path, db_manager, socketio, config={})
            # Load saved ROI points from database
            try:
                saved_roi = db_manager.get_channel_config(channel_id, 'RestrictedAreaMonitor', 'roi')
                logger.info(f"📐 Loading ROI from database for {channel_id}: type={type(saved_roi)}, data={saved_roi}")
                if saved_roi:
                    # Handle both formats: list or dict with 'main' key
                    if isinstance(saved_roi, dict) and 'main' in saved_roi:
                        module.set_roi_points(saved_roi['main'])
                        logger.info(f"✅ Loaded saved ROI (dict format) for RestrictedAreaMonitor on channel {channel_id}: {len(saved_roi['main'])} points")
                    elif isinstance(saved_roi, list):
                        module.set_roi_points(saved_roi)
                        logger.info(f"✅ Loaded saved ROI (list format) for RestrictedAreaMonitor on channel {channel_id}: {len(saved_roi)} points")
                    else:
                        logger.warning(f"⚠️ Unknown ROI format for {channel_id}: {type(saved_roi)}")
                else:
                    logger.info(f"ℹ️ No saved ROI found for RestrictedAreaMonitor on channel {channel_id}")
            except Exception as e:
                logger.error(f"❌ Could not load saved ROI for RestrictedAreaMonitor: {e}", exc_info=True)
        elif app_name == 'DressCodeMonitoring':
            module = DressCodeMonitoring(channel_id, socketio, db_manager, app)
        elif app_name == 'PPEMonitoring':
            module = PPEMonitoring(channel_id, socketio, db_manager, app)
        elif app_name == 'CrowdDetection':
            module = CrowdDetection(channel_id, socketio, db_manager, app)
        else:
            return jsonify({'success': False, 'error': 'Unknown app type'})
        
        # Add module to processor
        processor.add_module(app_name, module)
        channel_modules[channel_id][app_name] = module
        
        # Update config
        if channel_id not in app_configs[app_name]['channels']:
            app_configs[app_name]['channels'][channel_id] = {
                'name': f"Channel {channel_id}",
                'status': 'online',
                'video_source': video_source,
                'source_type': source_type,
                'shared': True,
                'active_modules': []
            }
        
        # Add this module to active modules list
        if app_name not in app_configs[app_name]['channels'][channel_id].get('active_modules', []):
            app_configs[app_name]['channels'][channel_id].setdefault('active_modules', []).append(app_name)
        
        # Update other app configs to show shared status
        for other_app in app_configs:
            if other_app != app_name and channel_id in channel_modules[channel_id]:
                if channel_id not in app_configs[other_app]['channels']:
                    app_configs[other_app]['channels'][channel_id] = {
                        'name': f"Channel {channel_id}",
                        'status': 'online',
                        'video_source': video_source,
                        'source_type': source_type,
                        'shared': True,
                        'active_modules': []
                    }
                # Update active modules for other apps
                current_modules = list(channel_modules[channel_id].keys())
                app_configs[other_app]['channels'][channel_id]['active_modules'] = current_modules
        
        logger.info(f"Started {app_name} on channel {channel_id} ({source_type}: {video_source})")
        return jsonify({
            'success': True, 
            'shared': True, 
            'active_modules': list(channel_modules[channel_id].keys()),
            'source_type': source_type,
            'video_source': video_source
        })
        
    except Exception as e:
        logger.error(f"Error starting channel: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/stop_channel', methods=['POST'])
@login_required
def stop_channel():
    """Stop processing a video channel or remove a module from shared channel"""
    data = request.json
    app_name = data.get('app_name')
    channel_id = data.get('channel_id')
    
    try:
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            
            # Remove the specific module
            if app_name in channel_modules.get(channel_id, {}):
                processor.remove_module(app_name)
                del channel_modules[channel_id][app_name]
                
                # Update app config
                if channel_id in app_configs[app_name]['channels']:
                    app_configs[app_name]['channels'][channel_id]['status'] = 'offline'
                
                # If no modules left, stop the entire processor
                if not channel_modules[channel_id]:
                    processor.stop()
                    del shared_video_processors[channel_id]
                    del channel_modules[channel_id]
                    
                    # Update all app configs
                    for app_config in app_configs.values():
                        if channel_id in app_config['channels']:
                            app_config['channels'][channel_id]['status'] = 'offline'
                else:
                    # Update active modules list for all apps
                    remaining_modules = list(channel_modules[channel_id].keys())
                    for app_config in app_configs.values():
                        if channel_id in app_config['channels']:
                            app_config['channels'][channel_id]['active_modules'] = remaining_modules
        
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error stopping channel: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_roi', methods=['POST'])
def set_roi():
    """Set ROI for queue monitoring on shared channel"""
    data = request.json
    app_name = data.get('app_name')
    channel_id = data.get('channel_id')
    roi_points = data.get('roi_points')
    
    try:
        logger.info(f"Setting ROI for {app_name} on channel {channel_id}")
        logger.info(f"ROI Points: {json.dumps(roi_points, indent=2)}")
        
        # Try shared processor first
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            # Access modules directly from processor.modules dictionary
            if hasattr(processor, 'modules') and app_name in processor.modules:
                module = processor.modules[app_name]
                if module and hasattr(module, 'set_roi'):
                    module.set_roi(roi_points)
                    logger.info(f"✓ ROI successfully set for {app_name} on channel {channel_id}")
                    return jsonify({'success': True})
        
        # Fallback to channel_modules
        if channel_id in channel_modules and app_name in channel_modules[channel_id]:
            module = channel_modules[channel_id][app_name]
            if hasattr(module, 'set_roi'):
                module.set_roi(roi_points)
                logger.info(f"✓ ROI successfully set for {app_name} on channel {channel_id}")
                return jsonify({'success': True})
        
        logger.warning(f"Module {app_name} not found on channel {channel_id}")
        logger.warning(f"Available modules in processor: {list(processor.modules.keys()) if channel_id in shared_video_processors and hasattr(shared_video_processors[channel_id], 'modules') else 'N/A'}")
        logger.warning(f"Available modules in channel_modules: {list(channel_modules[channel_id].keys()) if channel_id in channel_modules else 'N/A'}")
        return jsonify({'success': False, 'error': 'Module not found on this channel'})
    except Exception as e:
        logger.error(f"Error setting ROI: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_material_theft_roi/<channel_id>')
def get_material_theft_roi(channel_id):
    """Get current ROI configuration for MaterialTheftMonitor"""
    try:
        # Try shared processor first
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            # Access modules directly from processor.modules dictionary
            if hasattr(processor, 'modules') and 'MaterialTheftMonitor' in processor.modules:
                module = processor.modules['MaterialTheftMonitor']
                if module and hasattr(module, 'roi_points'):
                    roi_points = module.roi_points.tolist() if hasattr(module.roi_points, 'tolist') else module.roi_points
                    return jsonify({'success': True, 'roi_points': roi_points})
        
        # Fallback to channel_modules
        if channel_id in channel_modules and 'MaterialTheftMonitor' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['MaterialTheftMonitor']
            if hasattr(module, 'roi_points'):
                roi_points = module.roi_points.tolist() if hasattr(module.roi_points, 'tolist') else module.roi_points
                return jsonify({'success': True, 'roi_points': roi_points})
        
        return jsonify({'success': False, 'error': 'MaterialTheftMonitor not found on this channel'})
    except Exception as e:
        logger.error(f"Error getting MaterialTheftMonitor ROI: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/reset_material_theft_background/<channel_id>', methods=['POST'])
def reset_material_theft_background(channel_id):
    """Reset the background subtractor for MaterialTheftMonitor to relearn background"""
    try:
        # Try shared processor first
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if hasattr(processor, 'modules') and 'MaterialTheftMonitor' in processor.modules:
                module = processor.modules['MaterialTheftMonitor']
                if module and hasattr(module, 'reset_background'):
                    module.reset_background()
                    logger.info(f"✓ Background reset for MaterialTheftMonitor on channel {channel_id}")
                    return jsonify({'success': True, 'message': 'Background subtractor reset successfully'})
        
        # Fallback to channel_modules
        if channel_id in channel_modules and 'MaterialTheftMonitor' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['MaterialTheftMonitor']
            if hasattr(module, 'reset_background'):
                module.reset_background()
                logger.info(f"✓ Background reset for MaterialTheftMonitor on channel {channel_id}")
                return jsonify({'success': True, 'message': 'Background subtractor reset successfully'})
        
        return jsonify({'success': False, 'error': 'MaterialTheftMonitor not found on this channel'})
    except Exception as e:
        logger.error(f"Error resetting MaterialTheftMonitor background: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/reload_material_theft_roi/<channel_id>', methods=['POST'])
def reload_material_theft_roi(channel_id):
    """Reload ROI configuration from channels.json for MaterialTheftMonitor"""
    try:
        # Load config from channels.json
        config_path = Path('config/channels.json')
        if not config_path.exists():
            return jsonify({'success': False, 'error': 'channels.json not found'})
        
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        # Find the channel configuration
        channel_config = None
        for ch in config.get('channels', []):
            if ch.get('channel_id') == channel_id:
                channel_config = ch
                break
        
        if not channel_config:
            return jsonify({'success': False, 'error': f'Channel {channel_id} not found in config'})
        
        # Find MaterialTheftMonitor config
        mtm_config = None
        for module in channel_config.get('modules', []):
            if module.get('type') == 'MaterialTheftMonitor':
                mtm_config = module.get('config', {})
                break
        
        if not mtm_config or 'roi_points' not in mtm_config:
            return jsonify({'success': False, 'error': 'MaterialTheftMonitor ROI not found in config'})
        
        roi_points = mtm_config['roi_points']
        logger.info(f"Reloading ROI from config for {channel_id}: {roi_points}")
        
        # Update the module
        module_updated = False
        
        # Try shared processor first
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if hasattr(processor, 'modules') and 'MaterialTheftMonitor' in processor.modules:
                module = processor.modules['MaterialTheftMonitor']
                if module and hasattr(module, 'set_roi'):
                    module.set_roi(roi_points)
                    module_updated = True
                    logger.info(f"✓ ROI reloaded for MaterialTheftMonitor on channel {channel_id}")
        
        # Also update channel_modules
        if channel_id in channel_modules and 'MaterialTheftMonitor' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['MaterialTheftMonitor']
            if hasattr(module, 'set_roi'):
                module.set_roi(roi_points)
                module_updated = True
        
        if module_updated:
            return jsonify({'success': True, 'message': 'ROI reloaded from config', 'roi_points': roi_points})
        else:
            return jsonify({'success': False, 'error': 'MaterialTheftMonitor module not found'})
            
    except Exception as e:
        logger.error(f"Error reloading MaterialTheftMonitor ROI: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_roi/<app_name>/<channel_id>')
def get_roi(app_name, channel_id):
    """Get current ROI configuration for queue monitoring"""
    try:
        if channel_id in channel_modules and app_name in channel_modules[channel_id]:
            module = channel_modules[channel_id][app_name]
            if hasattr(module, 'get_roi'):
                roi_config = module.get_roi()
                return jsonify({'success': True, 'roi_config': roi_config})
        
        return jsonify({'success': False, 'error': 'Module not found on this channel'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_counting_line', methods=['POST'])
def set_counting_line():
    """Set counting line for people counter on shared channel"""
    data = request.json
    app_name = data.get('app_name')
    channel_id = data.get('channel_id')
    line_config = data.get('line_config')
    
    try:
        logger.info(f"Setting counting line for {app_name} on channel {channel_id}")
        logger.info(f"Counting Line Config: {json.dumps(line_config, indent=2)}")
        
        if channel_id in channel_modules and app_name in channel_modules[channel_id]:
            module = channel_modules[channel_id][app_name]
            if hasattr(module, 'set_counting_line'):
                module.set_counting_line(line_config)
                logger.info(f"✓ Counting line successfully set for {app_name} on channel {channel_id}")
                return jsonify({'success': True})
        
        logger.warning(f"Module {app_name} not found on channel {channel_id}")
        return jsonify({'success': False, 'error': 'Module not found on this channel'})
    except Exception as e:
        logger.error(f"Error setting counting line: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_counting_line/<app_name>/<channel_id>')
def get_counting_line(app_name, channel_id):
    """Get current counting line configuration for people counter"""
    try:
        if channel_id in channel_modules and app_name in channel_modules[channel_id]:
            module = channel_modules[channel_id][app_name]
            if hasattr(module, 'get_counting_line'):
                line_config = module.get_counting_line()
                return jsonify({'success': True, 'line_config': line_config})
        
        return jsonify({'success': False, 'error': 'Module not found on this channel'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_bag_settings', methods=['POST'])
def set_bag_settings():
    """Set bag detection settings (time threshold, proximity threshold, confidence)"""
    data = request.json
    app_name = 'BagDetection'
    channel_id = data.get('channel_id')
    time_threshold = data.get('time_threshold')
    proximity_threshold = data.get('proximity_threshold')
    confidence = data.get('confidence')
    alert_cooldown = data.get('alert_cooldown')
    
    try:
        if channel_id in channel_modules and app_name in channel_modules[channel_id]:
            module = channel_modules[channel_id][app_name]
            
            # Update module configuration
            if time_threshold is not None:
                module.time_threshold = float(time_threshold)
            if proximity_threshold is not None:
                module.proximity_threshold = float(proximity_threshold)
            if confidence is not None:
                module.confidence = float(confidence)
            if alert_cooldown is not None:
                module.alert_cooldown = float(alert_cooldown)
            
            return jsonify({
                'success': True,
                'message': 'Bag detection settings updated successfully',
                'settings': {
                    'time_threshold': module.time_threshold,
                    'proximity_threshold': module.proximity_threshold,
                    'confidence': module.confidence,
                    'alert_cooldown': module.alert_cooldown
                }
            })
        
        return jsonify({'success': False, 'error': 'Bag detection module not running on this channel'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_channel_status/<channel_id>')
def get_channel_status(channel_id):
    """Get status of all modules running on a channel"""
    try:
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            status = processor.get_status()
            
            # Add module-specific information
            module_info = {}
            for module_name, module in channel_modules.get(channel_id, {}).items():
                if hasattr(module, 'get_current_counts'):
                    module_info[module_name] = module.get_current_counts()
                elif hasattr(module, 'get_current_status'):
                    module_info[module_name] = module.get_current_status()
            
            status['module_info'] = module_info
            return jsonify(status)
        else:
            return jsonify({'error': 'Channel not found', 'channel_id': channel_id})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_module_analytics/<module_name>')
@login_required
def get_module_analytics(module_name):
    """Get analytics summary for a specific module"""
    try:
        analytics = {}
        
        if module_name == 'PeopleCounter':
            # Get today's IN/OUT counts from database (not in-memory counter)
            total_in = 0
            total_out = 0
            active_channels = []
            daily_data = []
            
            # Get list of active channels
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'PeopleCounter' in processor.get_active_modules():
                    active_channels.append(channel_id)
            
            # Get today's counts from database
            try:
                with app.app_context():
                    today_data = db_manager.get_today_footfall_count()
                    total_in = today_data.get('total_in', 0)
                    total_out = today_data.get('total_out', 0)
            except Exception as e:
                logger.error(f"Error getting today's footfall count: {e}")
                total_in = 0
                total_out = 0
            
            # Get daily footfall data for the last 7 days
            try:
                with app.app_context():
                    # Aggregate data from all active channels
                    from datetime import datetime, timedelta
                    end_date = datetime.now().date()
                    start_date = end_date - timedelta(days=6)
                    
                    # Create a dictionary to aggregate daily counts
                    daily_aggregated = {}
                    
                    for channel_id in active_channels:
                        report = db_manager.get_footfall_report(channel_id, period='7days')
                        for day_data in report.get('data', []):
                            date = day_data['date']
                            if date not in daily_aggregated:
                                daily_aggregated[date] = {'in_count': 0, 'out_count': 0}
                            daily_aggregated[date]['in_count'] += day_data['in_count']
                            daily_aggregated[date]['out_count'] += day_data['out_count']
                    
                    # Convert to sorted list
                    daily_data = [
                        {
                            'date': date,
                            'in_count': counts['in_count'],
                            'out_count': counts['out_count'],
                            'total': counts['in_count'] + counts['out_count']
                        }
                        for date, counts in sorted(daily_aggregated.items())
                    ]
            except Exception as e:
                logger.error(f"Error getting daily footfall data: {e}")
                daily_data = []
            
            # Get peak hour data
            peak_hour_info = {'peak_hour': 'N/A', 'traffic_count': 0}
            try:
                with app.app_context():
                    peak_hour_info = db_manager.get_people_counter_peak_hour()
            except Exception as e:
                logger.error(f"Error getting peak hour data: {e}")
            
            analytics = {
                'module': 'People Counter',
                'total_in': total_in,
                'total_out': total_out,
                'net_count': total_in - total_out,
                'active_channels': len(active_channels),
                'channels': active_channels,
                'daily_data': daily_data,  # Add daily breakdown
                'peak_hour': peak_hour_info.get('peak_hour', 'N/A'),
                'peak_hour_traffic': peak_hour_info.get('traffic_count', 0)
            }
            
        elif module_name == 'QueueMonitor':
            # Get total alerts and current queue stats
            total_alerts = 0
            current_queue_total = 0
            current_counter_total = 0
            active_channels = []
            
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'QueueMonitor' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'QueueMonitor' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['QueueMonitor']
                        status = module.get_status() if hasattr(module, 'get_status') else {}
                        # Get queue and counter counts from status
                        current_queue_total += status.get('queue_count', 0)
                        current_counter_total += status.get('counter_count', 0)
            
            # Get alert count from database (queue_alert type)
            try:
                with app.app_context():
                    alert_count = db_manager.get_alert_count('queue_alert', days=7)
                    total_alerts = alert_count if alert_count is not None else 0
            except Exception as e:
                logger.error(f"Error getting queue alert count: {e}")
                total_alerts = 0
            
            # Get queue violation count from database
            violation_count = 0
            try:
                with app.app_context():
                    violations = db_manager.get_queue_violations(limit=1000)
                    violation_count = len(violations) if violations else 0
            except Exception as e:
                logger.error(f"Error getting queue violations: {e}")
                violation_count = 0
            
            analytics = {
                'module': 'Queue Monitor',
                'total_alerts_7days': total_alerts,
                'total_violations': violation_count,
                'current_queue_count': current_queue_total,
                'current_counter_count': current_counter_total,
                'active_channels': len(active_channels),
                'channels': active_channels
            }
            
        elif module_name == 'BagDetection':
            # Get comprehensive bag detection analytics
            total_alerts = 0
            active_channels = []
            channel_details = []
            current_bags_tracked = 0
            current_unattended = 0
            
            # Get real-time data from active processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'BagDetection' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'BagDetection' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['BagDetection']
                        stats = module.get_statistics() if hasattr(module, 'get_statistics') else {}
                        current_bags_tracked += stats.get('bags_tracked', 0)
                        current_unattended += stats.get('current_unattended_bags', 0)
                        
                        channel_details.append({
                            'channel_id': channel_id,
                            'bags_tracked': stats.get('bags_tracked', 0),
                            'active_alerts': stats.get('active_alerts', 0),
                            'total_alerts': stats.get('total_alerts_triggered', 0),
                            'longest_unattended': stats.get('longest_unattended_time', 0),
                            'peak_bags': stats.get('peak_bags_count', 0)
                        })
            
            # Get historical analytics from database
            db_analytics = {}
            try:
                with app.app_context():
                    db_analytics = db_manager.get_bag_detection_analytics(days=7)
                    total_alerts = db_analytics.get('total_alerts', 0)
            except Exception as e:
                logger.error(f"Error getting bag detection analytics: {e}")
            
            # Sort channels by alert count (most risky first)
            channel_details.sort(key=lambda x: x['total_alerts'], reverse=True)
            
            # Identify most risky zone
            most_risky_now = channel_details[0] if channel_details else None
            most_risky_historical = db_analytics.get('most_risky_channel')
            
            analytics = {
                'module': 'Bag Detection',
                'total_alerts_7days': total_alerts,
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_bags_tracked': current_bags_tracked,
                'current_unattended_bags': current_unattended,
                'channel_details': channel_details,
                'most_risky_now': most_risky_now,
                'historical_data': {
                    'total_alerts': total_alerts,
                    'most_risky_channel': most_risky_historical,
                    'daily_trend': db_analytics.get('daily_trend', []),
                    'channels': db_analytics.get('channels', []),
                    'period_days': 7
                }
            }
            
        elif module_name == 'Heatmap':
            active_channels = []
            total_hotspots = 0
            channel_details = []
            
            # Get real-time data from active processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'HeatmapProcessor' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'HeatmapProcessor' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['HeatmapProcessor']
                        status = module.get_status() if hasattr(module, 'get_status') else {}
                        current_hotspots = status.get('hotspot_count', 0)
                        total_hotspots += current_hotspots
                        
                        channel_details.append({
                            'channel_id': channel_id,
                            'current_hotspots': current_hotspots,
                            'peak_hotspots': status.get('peak_hotspot_count', 0),
                            'peak_person_count': status.get('peak_person_count', 0),
                            'active_cells': status.get('active_cells', 0)
                        })
            
            # Sort channels by current hotspots (most crowded first)
            channel_details.sort(key=lambda x: x['current_hotspots'], reverse=True)
            
            # Get historical analytics from database
            db_analytics = {}
            try:
                with app.app_context():
                    db_analytics = db_manager.get_heatmap_analytics(days=7)
            except Exception as e:
                logger.error(f"Error getting heatmap analytics: {e}")
            
            # Identify most crowded zone
            most_crowded_now = channel_details[0] if channel_details else None
            most_crowded_historical = db_analytics.get('most_crowded_channel')
            
            analytics = {
                'module': 'Heatmap',
                'current_hotspots': total_hotspots,
                'active_channels': len(active_channels),
                'channels': active_channels,
                'channel_details': channel_details,
                'most_crowded_now': most_crowded_now,
                'historical_data': {
                    'total_snapshots': db_analytics.get('total_snapshots', 0),
                    'total_hotspots': db_analytics.get('total_hotspots', 0),
                    'most_crowded_channel': most_crowded_historical,
                    'period_days': 7,
                    'channels': db_analytics.get('channels', [])
                }
            }

            
        elif module_name == 'CashDetection':
            # Enhanced cash detection analytics
            active_channels = []
            channel_details = []
            current_detections_total = 0
            
            # Get real-time data from active processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'CashDetection' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'CashDetection' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['CashDetection']
                        stats = module.get_statistics() if hasattr(module, 'get_statistics') else {}
                        current_detections_total += stats.get('current_detections', 0)
                        
                        channel_details.append({
                            'channel_id': channel_id,
                            'current_detections': stats.get('current_detections', 0),
                            'total_alerts': stats.get('total_alerts', 0),
                            'total_detections': stats.get('total_detections', 0),
                            'peak_detections': stats.get('peak_detections', 0),
                            'detection_sessions': stats.get('detection_sessions', 0),
                            'avg_confidence': stats.get('avg_confidence', 0),
                            'highest_confidence': stats.get('highest_confidence', 0)
                        })
            
            # Get historical analytics from database
            db_analytics = {}
            try:
                with app.app_context():
                    db_analytics = db_manager.get_cash_detection_analytics(days=7)
            except Exception as e:
                logger.error(f"Error getting cash detection analytics: {e}")
            
            # Sort channels by snapshot count (most active first)
            channel_details.sort(key=lambda x: x['total_detections'], reverse=True)
            
            # Identify most active zone
            most_active_now = channel_details[0] if channel_details else None
            most_active_historical = db_analytics.get('most_active_channel')
            peak_hour = db_analytics.get('peak_hour')
            
            analytics = {
                'module': 'Cash Detection',
                'total_alerts_7days': db_analytics.get('total_snapshots', 0),
                'total_detections_7days': db_analytics.get('total_detections', 0),
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_detections': current_detections_total,
                'channel_details': channel_details,
                'most_active_now': most_active_now,
                'peak_hour': peak_hour,
                'historical_data': {
                    'total_snapshots': db_analytics.get('total_snapshots', 0),
                    'total_detections': db_analytics.get('total_detections', 0),
                    'most_active_channel': most_active_historical,
                    'daily_trend': db_analytics.get('daily_trend', []),
                    'hourly_distribution': db_analytics.get('hourly_distribution', []),
                    'channels': db_analytics.get('channels', []),
                    'period_days': 7
                }
            }
            
        elif module_name == 'FallDetection':
            active_channels = []
            current_falls = 0
            persons_tracked = 0
            
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'FallDetection' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    # Get current statistics from module if available
                    if channel_id in channel_modules and 'FallDetection' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['FallDetection']
                        stats = module.get_statistics() if hasattr(module, 'get_statistics') else {}
                        current_falls += stats.get('current_falls', 0)
                        persons_tracked += stats.get('persons_tracked', 0)
            
            # Also check configured channels (even if processor not running)
            for channel_id, modules_dict in channel_modules.items():
                if 'FallDetection' in modules_dict and channel_id not in active_channels:
                    active_channels.append(channel_id)
            
            # Get comprehensive analytics from database
            db_analytics = {}
            try:
                with app.app_context():
                    db_analytics = db_manager.get_fall_detection_analytics(days=7)
            except Exception as e:
                logger.error(f"Error getting fall detection analytics: {e}")
            
            # Get channel details
            channel_details = []
            most_risky_now = None
            
            for ch_data in db_analytics.get('channels', []):
                channel_details.append(ch_data)
                if most_risky_now is None or ch_data['fall_count'] > most_risky_now.get('fall_count', 0):
                    most_risky_now = ch_data
            
            analytics = {
                'module': 'Fall Detection',
                'total_falls_7days': db_analytics.get('total_falls', 0),
                'today_falls': db_analytics.get('today_falls', 0),
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_falls': current_falls,
                'persons_tracked': persons_tracked,
                'avg_fall_duration': db_analytics.get('avg_fall_duration', 0),
                'max_fall_duration': db_analytics.get('max_fall_duration', 0),
                'peak_hour': db_analytics.get('peak_hour', 'N/A'),
                'peak_hour_count': db_analytics.get('peak_hour_count', 0),
                'channel_details': channel_details,
                'most_risky_channel': most_risky_now,
                'response_categories': db_analytics.get('response_categories', {}),
                'historical_data': {
                    'daily_trend': db_analytics.get('daily_trend', []),
                    'most_incidents_channel': db_analytics.get('most_incidents_channel'),
                    'period_days': 7
                }
            }
        
        elif module_name == 'MoppingDetection':
            # Enhanced mopping detection analytics
            active_channels = []
            channel_details = []
            current_detections_total = 0
            
            # Get real-time data from active processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'MoppingDetection' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'MoppingDetection' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['MoppingDetection']
                        current_detections = getattr(module, 'detection_count', 0)
                        current_detections_total += current_detections
                        
                        # Get channel-specific stats from database
                        try:
                            with app.app_context():
                                ch_stats = db_manager.get_mopping_statistics(channel_id=channel_id, days=7)
                                channel_details.append({
                                    'channel_id': channel_id,
                                    'total_alerts': ch_stats.get('total_alerts', 0),
                                    'total_detections': ch_stats.get('total_detections', 0),
                                    'current_detections': current_detections
                                })
                        except Exception as e:
                            logger.error(f"Error getting channel stats for {channel_id}: {e}")
            
            # Get comprehensive analytics from database
            db_stats = {}
            most_active_channel = None
            try:
                with app.app_context():
                    db_stats = db_manager.get_mopping_statistics(days=7)
                    
                    # Find most active channel from channel_details
                    if channel_details:
                        most_active_channel = max(channel_details, key=lambda x: x['total_alerts'])
            except Exception as e:
                logger.error(f"Error getting mopping detection statistics: {e}")
            
            # Sort channels by alert count (most active first)
            channel_details.sort(key=lambda x: x['total_alerts'], reverse=True)
            
            # Get today's alerts
            today_alerts = 0
            try:
                with app.app_context():
                    today_stats = db_manager.get_mopping_statistics(days=1)
                    today_alerts = today_stats.get('total_alerts', 0)
            except Exception as e:
                logger.error(f"Error getting today's mopping stats: {e}")
            
            analytics = {
                'module': 'Mopping Detection',
                'total_alerts_7days': db_stats.get('total_alerts', 0),
                'today_alerts': today_alerts,
                'total_detections_7days': db_stats.get('total_detections', 0),
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_detections': current_detections_total,
                'channel_details': channel_details,
                'most_active_channel': most_active_channel,
                'peak_hour': 'N/A',  # Can be enhanced later
                'historical_data': {
                    'total_snapshots': db_stats.get('total_alerts', 0),
                    'most_active_channel': most_active_channel,
                    'daily_counts': db_stats.get('daily_counts', {}),
                    'period_days': 7
                }
            }
        
        elif module_name == 'SmokingDetection':
            # Enhanced smoking detection analytics
            active_channels = []
            channel_details = []
            current_detections_total = 0
            
            # Get real-time data from active processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'SmokingDetection' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'SmokingDetection' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['SmokingDetection']
                        current_detections = getattr(module, 'detection_count', 0)
                        current_detections_total += current_detections
                        
                        # Get channel-specific stats from database
                        try:
                            with app.app_context():
                                ch_stats = db_manager.get_smoking_statistics(channel_id=channel_id, days=7)
                                channel_details.append({
                                    'channel_id': channel_id,
                                    'total_alerts': ch_stats.get('total_alerts', 0),
                                    'total_detections': ch_stats.get('total_detections', 0),
                                    'current_detections': current_detections
                                })
                        except Exception as e:
                            logger.error(f"Error getting channel stats for {channel_id}: {e}")
            
            # Get comprehensive analytics from database
            db_stats = {}
            most_active_channel = None
            try:
                with app.app_context():
                    db_stats = db_manager.get_smoking_statistics(days=7)
                    
                    # Find most active channel from channel_details
                    if channel_details:
                        most_active_channel = max(channel_details, key=lambda x: x['total_alerts'])
            except Exception as e:
                logger.error(f"Error getting smoking detection statistics: {e}")
            
            # Sort channels by alert count (most active first)
            channel_details.sort(key=lambda x: x['total_alerts'], reverse=True)
            
            # Get today's alerts
            today_alerts = 0
            try:
                with app.app_context():
                    today_stats = db_manager.get_smoking_statistics(days=1)
                    today_alerts = today_stats.get('total_alerts', 0)
            except Exception as e:
                logger.error(f"Error getting today's smoking stats: {e}")
            
            analytics = {
                'module': 'Smoke & Fire Detection',
                'total_alerts_7days': db_stats.get('total_alerts', 0),
                'today_alerts': today_alerts,
                'total_detections_7days': db_stats.get('total_detections', 0),
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_detections': current_detections_total,
                'channel_details': channel_details,
                'most_active_channel': most_active_channel,
                'peak_hour': 'N/A',  # Can be enhanced later
                'historical_data': {
                    'total_snapshots': db_stats.get('total_alerts', 0),
                    'most_active_channel': most_active_channel,
                    'daily_counts': db_stats.get('daily_counts', {}),
                    'period_days': 7
                }
            }
        
        elif module_name == 'PersonSmokingDetection':
            # Smoking Detection analytics (cigarette smoking)
            active_channels = []
            channel_details = []
            current_detections_total = 0
            
            # Get real-time data from active processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'PersonSmokingDetection' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'PersonSmokingDetection' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['PersonSmokingDetection']
                        current_detections = getattr(module, 'detection_count', 0)
                        current_detections_total += current_detections
                        
                        # Get channel-specific stats from database
                        try:
                            with app.app_context():
                                alert_count = db_manager.get_alert_count('person_smoking_alert', days=7, channel_id=channel_id)
                                channel_details.append({
                                    'channel_id': channel_id,
                                    'total_alerts': alert_count if alert_count is not None else 0,
                                    'current_detections': current_detections
                                })
                        except Exception as e:
                            logger.error(f"Error getting channel stats for {channel_id}: {e}")
            
            # Get alert count from database
            total_alerts = 0
            try:
                with app.app_context():
                    alert_count = db_manager.get_alert_count('person_smoking_alert', days=7)
                    total_alerts = alert_count if alert_count is not None else 0
            except Exception as e:
                logger.error(f"Error getting person smoking alert count: {e}")
                total_alerts = 0
            
            # Get today's alerts
            today_alerts = 0
            try:
                with app.app_context():
                    today_alert_count = db_manager.get_alert_count('person_smoking_alert', days=1)
                    today_alerts = today_alert_count if today_alert_count is not None else 0
            except Exception as e:
                logger.error(f"Error getting today's person smoking alerts: {e}")
            
            # Sort channels by alert count (most active first)
            channel_details.sort(key=lambda x: x['total_alerts'], reverse=True)
            most_active_channel = channel_details[0] if channel_details else None
            
            analytics = {
                'module': 'Smoking Detection',
                'total_alerts_7days': total_alerts,
                'today_alerts': today_alerts,
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_detections': current_detections_total,
                'channel_details': channel_details,
                'most_active_channel': most_active_channel,
                'historical_data': {
                    'total_alerts': total_alerts,
                    'most_active_channel': most_active_channel,
                    'period_days': 7
                }
            }
        
        elif module_name == 'PhoneUsageDetection':
            # Enhanced phone usage detection analytics
            active_channels = []
            channel_details = []
            current_detections_total = 0
            
            # Get real-time data from active processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'PhoneUsageDetection' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'PhoneUsageDetection' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['PhoneUsageDetection']
                        current_detections = getattr(module, 'detection_count', 0)
                        current_detections_total += current_detections
                        
                        # Get channel-specific stats from database
                        try:
                            with app.app_context():
                                ch_stats = db_manager.get_phone_statistics(channel_id=channel_id, days=7)
                                channel_details.append({
                                    'channel_id': channel_id,
                                    'total_alerts': ch_stats.get('total_alerts', 0),
                                    'total_detections': ch_stats.get('total_detections', 0),
                                    'current_detections': current_detections
                                })
                        except Exception as e:
                            logger.error(f"Error getting channel stats for {channel_id}: {e}")
            
            # Get comprehensive analytics from database
            db_stats = {}
            most_active_channel = None
            try:
                with app.app_context():
                    db_stats = db_manager.get_phone_statistics(days=7)
                    
                    # Find most active channel from channel_details
                    if channel_details:
                        most_active_channel = max(channel_details, key=lambda x: x['total_alerts'])
            except Exception as e:
                logger.error(f"Error getting phone usage detection statistics: {e}")
            
            # Sort channels by alert count (most active first)
            channel_details.sort(key=lambda x: x['total_alerts'], reverse=True)
            
            # Get today's alerts
            today_alerts = 0
            try:
                with app.app_context():
                    today_stats = db_manager.get_phone_statistics(days=1)
                    today_alerts = today_stats.get('total_alerts', 0)
            except Exception as e:
                logger.error(f"Error getting today's phone stats: {e}")
            
            analytics = {
                'module': 'Phone Usage Detection',
                'total_alerts_7days': db_stats.get('total_alerts', 0),
                'today_alerts': today_alerts,
                'total_detections_7days': db_stats.get('total_detections', 0),
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_detections': current_detections_total,
                'channel_details': channel_details,
                'most_active_channel': most_active_channel,
                'peak_hour': 'N/A',  # Can be enhanced later
                'historical_data': {
                    'total_snapshots': db_stats.get('total_alerts', 0),
                    'most_active_channel': most_active_channel,
                    'daily_counts': db_stats.get('daily_counts', {}),
                    'period_days': 7
                }
            }
        
        elif module_name == 'RestrictedAreaMonitor':
            # Restricted Area Monitor analytics
            active_channels = []
            channel_details = []
            current_violations_total = 0
            
            # Get real-time data from active processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'RestrictedAreaMonitor' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'RestrictedAreaMonitor' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['RestrictedAreaMonitor']
                        current_violations = module.stats.get('total_violations', 0)
                        current_violations_total += current_violations
                        
                        # Get channel-specific stats from database
                        try:
                            with app.app_context():
                                ch_stats = db_manager.get_restricted_area_statistics(channel_id=channel_id, days=7)
                                channel_details.append({
                                    'channel_id': channel_id,
                                    'total_alerts': ch_stats.get('total_alerts', 0),
                                    'total_violations': ch_stats.get('total_violations', 0),
                                    'roi_defined': len(module.roi_points) >= 3
                                })
                        except Exception as e:
                            logger.error(f"Error getting restricted area stats for {channel_id}: {e}")
            
            # Get aggregated database statistics
            db_stats = {}
            most_active_channel = None
            try:
                with app.app_context():
                    db_stats = db_manager.get_restricted_area_statistics(days=7)
                    
                    # Find most active channel from channel_details
                    if channel_details:
                        most_active_channel = max(channel_details, key=lambda x: x['total_alerts'])
            except Exception as e:
                logger.error(f"Error getting restricted area statistics: {e}")
            
            # Sort channels by alert count (most active first)
            channel_details.sort(key=lambda x: x['total_alerts'], reverse=True)
            
            # Get today's alerts
            today_alerts = 0
            try:
                with app.app_context():
                    today_stats = db_manager.get_restricted_area_statistics(days=1)
                    today_alerts = today_stats.get('total_alerts', 0)
            except Exception as e:
                logger.error(f"Error getting today's restricted area stats: {e}")
            
            analytics = {
                'module': 'Restricted Area Monitor',
                'total_alerts_7days': db_stats.get('total_alerts', 0),
                'today_alerts': today_alerts,
                'total_violations_7days': db_stats.get('total_violations', 0),
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_violations': current_violations_total,
                'channel_details': channel_details,
                'most_active_channel': most_active_channel,
                'historical_data': {
                    'total_snapshots': db_stats.get('total_alerts', 0),
                    'most_active_channel': most_active_channel,
                    'daily_counts': db_stats.get('daily_counts', {}),
                    'period_days': 7
                }
            }
        
        elif module_name == 'UnauthorizedEntryMonitor':
            # Unauthorized Entry Monitor analytics
            active_channels = []
            channel_details = []
            current_detections_total = 0
            total_alerts = 0
            
            # Get real-time data from active processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'UnauthorizedEntryMonitor' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'UnauthorizedEntryMonitor' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['UnauthorizedEntryMonitor']
                        status = module.get_status() if hasattr(module, 'get_status') else {}
                        current_detections = status.get('current_detections', 0)
                        current_detections_total += current_detections
                        
                        channel_details.append({
                            'channel_id': channel_id,
                            'current_detections': current_detections,
                            'total_alerts': status.get('total_alerts', 0),
                            'total_detections': status.get('total_detections', 0),
                            'peak_detections': status.get('peak_detections', 0),
                            'detection_sessions': status.get('detection_sessions', 0)
                        })
            
            # Get alert count from database
            try:
                with app.app_context():
                    alert_count = db_manager.get_alert_count('unauthorized_entry_alert', days=7)
                    total_alerts = alert_count if alert_count is not None else 0
            except Exception as e:
                logger.error(f"Error getting unauthorized entry alert count: {e}")
                total_alerts = 0
            
            # Get today's alerts
            today_alerts = 0
            try:
                with app.app_context():
                    today_alert_count = db_manager.get_alert_count('unauthorized_entry_alert', days=1)
                    today_alerts = today_alert_count if today_alert_count is not None else 0
            except Exception as e:
                logger.error(f"Error getting today's unauthorized entry alerts: {e}")
            
            # Sort channels by alert count (most active first)
            channel_details.sort(key=lambda x: x['total_alerts'], reverse=True)
            most_active_channel = channel_details[0] if channel_details else None
            
            analytics = {
                'module': 'Unauthorized Entry Monitor',
                'total_alerts_7days': total_alerts,
                'today_alerts': today_alerts,
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_detections': current_detections_total,
                'channel_details': channel_details,
                'most_active_channel': most_active_channel,
                'historical_data': {
                    'total_alerts': total_alerts,
                    'most_active_channel': most_active_channel,
                    'period_days': 7
                }
            }

        elif module_name == 'MaterialTheftMonitor':
            # Material Theft / Misuse analytics
            active_channels = []
            channel_details = []
            total_alerts = 0
            today_alerts = 0

            for channel_id, processor in shared_video_processors.items():
                # Check if MaterialTheftMonitor is in active modules
                if 'MaterialTheftMonitor' not in processor.get_active_modules():
                    continue
                
                # Check if processor is actually working (same logic as stream subscription)
                is_working = False
                if processor.is_running:
                    is_working = True
                else:
                    # Check if processor is actually working even if flag is False
                    thread_alive = hasattr(processor, 'processing_thread') and processor.processing_thread and processor.processing_thread.is_alive()
                    has_recent_frame = False
                    if hasattr(processor, 'latest_raw_frame') and processor.latest_raw_frame is not None:
                        has_recent_frame = True
                    elif hasattr(processor, 'latest_annotated_frame') and processor.latest_annotated_frame is not None:
                        has_recent_frame = True
                    is_working = thread_alive or has_recent_frame
                
                if is_working:
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'MaterialTheftMonitor' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['MaterialTheftMonitor']
                        status = module.get_status() if hasattr(module, 'get_status') else {}
                        channel_details.append({
                            'channel_id': channel_id,
                            'last_alert_time': status.get('last_alert_time'),
                            'still_counter': status.get('still_counter', 0),
                            'frame_count': status.get('frame_count', 0)
                        })

            try:
                with app.app_context():
                    total_alerts = db_manager.get_alert_count('material_theft_alert', days=7) or 0
                    today_alerts = db_manager.get_alert_count('material_theft_alert', days=1) or 0
            except Exception as e:
                logger.error(f"Error getting material theft alert count: {e}")

            analytics = {
                'module': 'Material Theft / Misuse',
                'total_alerts_7days': total_alerts,
                'today_alerts': today_alerts,
                'active_channels': len(active_channels),
                'channels': active_channels,
                'channel_details': channel_details,
                'most_active_channel': channel_details[0] if channel_details else None,
                'historical_data': {
                    'total_alerts': total_alerts,
                    'period_days': 7
                }
            }
        
        elif module_name == 'PPEMonitoring':
            try:
                with app.app_context():
                    total_alerts = db_manager.get_alert_count('ppe_alert', days=365)  # All time (last year)
                    today_alerts = db_manager.get_alert_count('ppe_alert', days=1)
                    
                    # Get active channels - check both running processors and configured modules
                    active_channels = []
                    total_violations = 0
                    
                    # First check running processors
                    for channel_id, processor in shared_video_processors.items():
                        if processor.is_running and 'PPEMonitoring' in processor.get_active_modules():
                            active_channels.append(channel_id)
                            # Get stats from module if available
                            if channel_id in channel_modules and 'PPEMonitoring' in channel_modules[channel_id]:
                                module = channel_modules[channel_id]['PPEMonitoring']
                                if hasattr(module, 'total_violations'):
                                    total_violations += module.total_violations
                    
                    # Also check configured modules (even if processor not running yet)
                    for channel_id, modules_dict in channel_modules.items():
                        if 'PPEMonitoring' in modules_dict and channel_id not in active_channels:
                            # Module is configured but processor might not be running
                            processor = shared_video_processors.get(channel_id)
                            if processor and processor.is_running:
                                # Should have been caught above, but double-check
                                if channel_id not in active_channels:
                                    active_channels.append(channel_id)
                            # Even if not running, we can still show it as configured
                            # (This helps show channels that are set up but not yet started)
                    
                    analytics = {
                        'module': 'PPE Compliance',
                        'total_alerts': total_alerts,
                        'today_alerts': today_alerts,
                        'active_channels': len(active_channels),
                        'total_violations': total_violations,
                        'channels': [{'channel_id': ch_id, 'status': 'active'} for ch_id in active_channels]
                    }
                    
                    return jsonify({'success': True, 'analytics': analytics})
            except Exception as e:
                logger.error(f"Error getting PPE analytics: {e}", exc_info=True)
                return jsonify({'success': False, 'error': str(e)})
        
        elif module_name == 'DressCodeMonitoring':
            active_channels = []
            total_violations = 0
            
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'DressCodeMonitoring' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    # Get stats from module if available
                    if channel_id in channel_modules and 'DressCodeMonitoring' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['DressCodeMonitoring']
                        if hasattr(module, 'total_violations'):
                            total_violations += module.total_violations
            
            # Get comprehensive statistics from database
            db_stats = {}
            try:
                with app.app_context():
                    db_stats = db_manager.get_dresscode_stats(days=7)
            except Exception as e:
                logger.error(f"Error getting dress code stats: {e}")
            
            analytics = {
                'module': 'Dress Code Monitoring',
                'total_violations': db_stats.get('total_violations', 0),
                'violation_types': db_stats.get('violation_types', {}),
                'uniform_colors': db_stats.get('uniform_colors', {}),
                'active_channels': len(active_channels),
                'channels': active_channels,
                'current_violations': total_violations,
                'period_days': 7
            }
            
        elif module_name == 'GroomingDetection':
            active_channels = []
            total_alerts = 0
            
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and module_name in processor.get_active_modules():
                    active_channels.append(channel_id)
        
        elif module_name == 'CrowdDetection':
            active_channels = []
            current_crowd_total = 0
            total_alerts = 0
            
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'CrowdDetection' in processor.get_active_modules():
                    active_channels.append(channel_id)
                    if channel_id in channel_modules and 'CrowdDetection' in channel_modules[channel_id]:
                        module = channel_modules[channel_id]['CrowdDetection']
                        status = module.get_status() if hasattr(module, 'get_status') else {}
                        current_crowd_total += status.get('crowd_count', 0)
            
            # Get alert count from database
            try:
                with app.app_context():
                    alert_count = db_manager.get_alert_count('crowd_alert', days=7)
                    total_alerts = alert_count if alert_count is not None else 0
            except Exception as e:
                logger.error(f"Error getting crowd alert count: {e}", exc_info=True)
                total_alerts = 0
            
            analytics = {
                'module': 'Crowd Detection',
                'total_alerts_7days': total_alerts,
                'current_crowd_count': current_crowd_total,
                'crowd_threshold': 5,  # Default threshold
                'active_channels': len(active_channels),
                'channels': active_channels
            }
        
        elif module_name == 'TableServiceMonitor':
            active_channels = []
            total_alerts = 0
            avg_unclean_time = 0
            max_unclean_time = 0
            avg_reset_time = 0
            max_reset_time = 0
            
            # Check running processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'TableServiceMonitor' in processor.get_active_modules():
                    active_channels.append(channel_id)
            
            # Also check configured channels (even if processor not running)
            for channel_id, modules_dict in channel_modules.items():
                if 'TableServiceMonitor' in modules_dict and channel_id not in active_channels:
                    active_channels.append(channel_id)
            
            # Get violations from database and calculate statistics
            try:
                with app.app_context():
                    from datetime import datetime, timedelta
                    import json
                    
                    date_threshold = datetime.now() - timedelta(days=7)
                    
                    # Get all table cleanliness violations from last 7 days
                    all_violations = db.session.query(db_manager.TableCleanlinessViolation).filter(
                        db_manager.TableCleanlinessViolation.created_at >= date_threshold
                    ).all()
                    
                    unclean_durations = []
                    reset_durations = []
                    
                    for violation in all_violations:
                        if violation.alert_data:
                            try:
                                alert_data = json.loads(violation.alert_data) if isinstance(violation.alert_data, str) else violation.alert_data
                                violation_type = violation.violation_type or alert_data.get('violation_type')
                                
                                # Extract duration based on violation type
                                if violation_type == 'unclean_table':
                                    unclean_duration = alert_data.get('unclean_duration')
                                    if unclean_duration is not None:
                                        try:
                                            dur = float(unclean_duration)
                                            if dur > 0:
                                                unclean_durations.append(dur)
                                        except (ValueError, TypeError):
                                            pass
                                
                                elif violation_type == 'slow_reset':
                                    reset_duration = alert_data.get('reset_duration')
                                    if reset_duration is not None:
                                        try:
                                            dur = float(reset_duration)
                                            if dur > 0:
                                                reset_durations.append(dur)
                                        except (ValueError, TypeError):
                                            pass
                                
                                total_alerts += 1
                            except Exception as e:
                                logger.debug(f"Error parsing alert_data for violation {violation.id}: {e}")
                                total_alerts += 1
                        else:
                            total_alerts += 1
                    
                    # Calculate average and max times for unclean tables
                    if unclean_durations:
                        # Filter out unrealistic values (cap at 1 hour = 3600 seconds)
                        reasonable_unclean = [d for d in unclean_durations if 0 < d <= 3600]
                        if reasonable_unclean:
                            max_unclean_time = max(reasonable_unclean)
                            # Trim top 5% for average calculation
                            if len(reasonable_unclean) >= 20:
                                sorted_dur = sorted(reasonable_unclean)
                                trim_n = max(1, int(len(sorted_dur) * 0.05))
                                trimmed = sorted_dur[:-trim_n]
                                avg_unclean_time = sum(trimmed) / len(trimmed) if trimmed else 0
                            else:
                                avg_unclean_time = sum(reasonable_unclean) / len(reasonable_unclean)
                    
                    # Calculate average and max times for slow reset
                    if reset_durations:
                        # Filter out unrealistic values (cap at 1 hour = 3600 seconds)
                        reasonable_reset = [d for d in reset_durations if 0 < d <= 3600]
                        if reasonable_reset:
                            max_reset_time = max(reasonable_reset)
                            # Trim top 5% for average calculation
                            if len(reasonable_reset) >= 20:
                                sorted_dur = sorted(reasonable_reset)
                                trim_n = max(1, int(len(sorted_dur) * 0.05))
                                trimmed = sorted_dur[:-trim_n]
                                avg_reset_time = sum(trimmed) / len(trimmed) if trimmed else 0
                            else:
                                avg_reset_time = sum(reasonable_reset) / len(reasonable_reset)
                    
            except Exception as e:
                logger.error(f"Error getting table cleanliness analytics: {e}")
                total_alerts = 0
            
            analytics = {
                'module': 'Table Cleanliness',
                'total_alerts_7days': total_alerts,
                'active_channels': len(active_channels),
                'channels': active_channels,
                'avg_unclean_time': round(avg_unclean_time, 1),  # Average time tables remain unclean (seconds)
                'max_unclean_time': round(max_unclean_time, 1),  # Maximum time a table remained unclean (seconds)
                'avg_reset_time': round(avg_reset_time, 1),  # Average reset time (seconds)
                'max_reset_time': round(max_reset_time, 1)  # Maximum reset time (seconds)
            }
        
        elif module_name == 'ServiceDisciplineMonitor':
            active_channels = []
            total_alerts = 0
            avg_wait_time = 0
            max_wait_time = 0
            avg_order_wait_time = 0
            max_order_wait_time = 0
            avg_service_wait_time = 0
            max_service_wait_time = 0
            
            # Check running processors
            for channel_id, processor in shared_video_processors.items():
                if processor.is_running and 'ServiceDisciplineMonitor' in processor.get_active_modules():
                    active_channels.append(channel_id)
            
            # Also check configured channels (even if processor not running)
            for channel_id, modules_dict in channel_modules.items():
                if 'ServiceDisciplineMonitor' in modules_dict and channel_id not in active_channels:
                    active_channels.append(channel_id)
            
            # Get violations from database
            try:
                with app.app_context():
                    from datetime import datetime, timedelta
                    from zoneinfo import ZoneInfo
                    import json
                    
                    # Use IST timezone to match database timestamps (which use get_ist_now())
                    ist_now = datetime.now(ZoneInfo("Asia/Kolkata"))
                    date_threshold = ist_now - timedelta(days=7)
                    logger.info(f"📊 Service Discipline Analytics: Querying violations from last 7 days (since {date_threshold} IST)")
                    
                    # Get all violations from last 7 days
                    # Use db_manager.TableServiceViolation instead of importing
                    # Only select columns that exist (order_wait_time and service_wait_time may not exist yet)
                    try:
                        # Try to query with all columns first
                        all_violations = db.session.query(db_manager.TableServiceViolation).filter(
                            db_manager.TableServiceViolation.created_at >= date_threshold
                        ).all()
                        logger.info(f"📊 Service Discipline Analytics: Found {len(all_violations)} total violations in database (last 7 days)")
                    except Exception as db_error:
                        # Rollback any failed transaction first
                        try:
                            db.session.rollback()
                        except Exception:
                            pass  # Ignore rollback errors
                        
                        # If columns don't exist, use raw SQL to query only existing columns
                        error_str = str(db_error)
                        if ('order_wait_time' in error_str or 'service_wait_time' in error_str or 
                            'UndefinedColumn' in error_str or 'InFailedSqlTransaction' in error_str):
                            logger.warning("order_wait_time/service_wait_time columns don't exist yet, using raw SQL query")
                            # Use raw SQL to query only columns that exist
                            try:
                                result = db.session.execute(
                                    db.text("""
                                        SELECT id, channel_id, table_id, waiting_time, 
                                               snapshot_filename, snapshot_path, alert_data, 
                                               file_size, created_at
                                        FROM table_service_violations
                                        WHERE created_at >= :date_threshold
                                    """),
                                    {'date_threshold': date_threshold}
                                )
                                
                                # Create wrapper objects with same interface
                                class ViolationWrapper:
                                    def __init__(self, row):
                                        self.id = row[0]
                                        self.channel_id = row[1]
                                        self.table_id = row[2]
                                        self.waiting_time = row[3]
                                        self.snapshot_filename = row[4]
                                        self.snapshot_path = row[5]
                                        self.alert_data = row[6]
                                        self.file_size = row[7]
                                        self.created_at = row[8]
                                        # These columns don't exist - will be extracted from alert_data
                                        self.order_wait_time = None
                                        self.service_wait_time = None
                                
                                all_violations = [ViolationWrapper(row) for row in result]
                            except Exception as fallback_error:
                                logger.error(f"Error in fallback SQL query: {fallback_error}")
                                # If fallback also fails, return empty results
                                all_violations = []
                        else:
                            logger.error(f"Database error getting service discipline analytics: {db_error}")
                            # Return empty results instead of raising to prevent dashboard errors
                            all_violations = []
                    
                    # Service discipline analytics:
                    # The reports UI currently treats any record with waiting_time > 0 as relevant.
                    # Some code paths store different keys in alert_data, so we extract robustly.
                    relevant_violations = []
                    wait_times = []
                    order_wait_times = []
                    service_wait_times = []
                    excluded_types = {'unclean_table', 'slow_reset', 'wrong_uniform'}
                    
                    for violation in all_violations:
                        # Robustly extract wait times from columns first (new fields)
                        wt = None
                        order_wt = None
                        service_wt = None
                        violation_type = None
                        
                        try:
                            # Check if columns exist before accessing (database migration may not be complete)
                            if hasattr(violation, 'order_wait_time') and violation.order_wait_time is not None:
                                order_wt = float(violation.order_wait_time)
                            if hasattr(violation, 'service_wait_time') and violation.service_wait_time is not None:
                                service_wt = float(violation.service_wait_time)
                            if violation.waiting_time is not None:
                                wt = float(violation.waiting_time)
                        except Exception:
                            pass
                        
                        # Parse alert_data if available
                        if violation.alert_data:
                            try:
                                alert_data = json.loads(violation.alert_data) if isinstance(violation.alert_data, str) else violation.alert_data
                                violation_type = alert_data.get('violation_type')
                                
                                # Skip non-service discipline types stored in the same table
                                if violation_type in excluded_types:
                                    continue
                                
                                # Fallback to alert_data if columns not set
                                if order_wt is None or service_wt is None:
                                    for k in ('order_wait_time', 'service_wait_time', 'waiting_time', 'wait_time'):
                                        if k in alert_data and alert_data[k] is not None:
                                            try:
                                                val = float(alert_data[k])
                                                if k == 'order_wait_time' and order_wt is None:
                                                    order_wt = val
                                                elif k == 'service_wait_time' and service_wt is None:
                                                    service_wt = val
                                                elif k in ('waiting_time', 'wait_time') and wt is None:
                                                    wt = val
                                            except Exception:
                                                pass
                            except:
                                pass
                        
                        # Include all records with wait times (both violations and completed orders)
                        # If no alert_data/violation_type, it's a completed order (saved via add_table_service_order)
                        if violation_type is None:
                            # Completed order - include if has wait times
                            if (order_wt is not None and order_wt > 0) or (service_wt is not None and service_wt > 0):
                                relevant_violations.append(violation)
                                if order_wt is not None and order_wt > 0:
                                    order_wait_times.append(order_wt)
                                if service_wt is not None and service_wt > 0:
                                    service_wait_times.append(service_wt)
                        else:
                            # Has violation_type - include if wait time > 0 OR explicitly service_discipline
                            if (wt is not None and wt > 0) or (order_wt is not None and order_wt > 0) or (service_wt is not None and service_wt > 0) or (violation_type == 'service_discipline'):
                                relevant_violations.append(violation)
                                if wt is not None:
                                    wait_times.append(wt)
                                if order_wt is not None and order_wt > 0:
                                    order_wait_times.append(order_wt)
                                if service_wt is not None and service_wt > 0:
                                    service_wait_times.append(service_wt)
                    
                    total_alerts = len(relevant_violations)
                    logger.info(f"📊 Service Discipline Analytics: {total_alerts} relevant violations after filtering (excluded types: {excluded_types})")
                    logger.info(f"📊 Wait times found: {len(wait_times)} total, {len(order_wait_times)} order, {len(service_wait_times)} service")
                    
                    # Calculate avg/max wait time with filtering to avoid bad/outlier values.
                    # Some records can be huge due to tracking resets/timeouts; keep a sane range.
                    max_reasonable_wait = 3600.0  # hard cap at 1 hour
                    reasonable_wait_times = [wt for wt in wait_times if 0 < wt <= max_reasonable_wait]
                    reasonable_order_wait_times = [wt for wt in order_wait_times if 0 < wt <= max_reasonable_wait]
                    reasonable_service_wait_times = [wt for wt in service_wait_times if 0 < wt <= max_reasonable_wait]
                    
                    # Trim the top 5% to reduce impact of extreme but "reasonable" values.
                    # This keeps avg realistic while still reflecting long waits.
                    def trim_and_avg(times_list):
                        if len(times_list) >= 20:
                            sorted_times = sorted(times_list)
                            trim_n = max(1, int(len(sorted_times) * 0.05))
                            trimmed = sorted_times[:-trim_n]  # drop top 5%
                            return trimmed
                        return times_list
                    
                    trimmed_wait_times = trim_and_avg(reasonable_wait_times)
                    trimmed_order_wait_times = trim_and_avg(reasonable_order_wait_times)
                    trimmed_service_wait_times = trim_and_avg(reasonable_service_wait_times)
                    
                    if reasonable_wait_times:
                        max_wait_time = max(reasonable_wait_times)
                    else:
                        max_wait_time = 0
                    
                    if trimmed_wait_times:
                        avg_wait_time = sum(trimmed_wait_times) / len(trimmed_wait_times)
                    else:
                        avg_wait_time = 0
                    
                    # Calculate order wait time stats
                    if reasonable_order_wait_times:
                        max_order_wait_time = max(reasonable_order_wait_times)
                    else:
                        max_order_wait_time = 0
                    
                    if trimmed_order_wait_times:
                        avg_order_wait_time = sum(trimmed_order_wait_times) / len(trimmed_order_wait_times)
                    else:
                        avg_order_wait_time = 0
                    
                    # Calculate service wait time stats
                    if reasonable_service_wait_times:
                        max_service_wait_time = max(reasonable_service_wait_times)
                    else:
                        max_service_wait_time = 0
                    
                    if trimmed_service_wait_times:
                        avg_service_wait_time = sum(trimmed_service_wait_times) / len(trimmed_service_wait_times)
                    else:
                        avg_service_wait_time = 0
            except Exception as e:
                logger.error(f"❌ Error getting service discipline analytics: {e}", exc_info=True)
                total_alerts = 0
                avg_wait_time = 0
                max_wait_time = 0
                avg_order_wait_time = 0
                max_order_wait_time = 0
                avg_service_wait_time = 0
                max_service_wait_time = 0
            
            analytics = {
                'module': 'Service Discipline',
                'total_alerts_7days': total_alerts,
                'avg_wait_time': round(avg_wait_time, 1),
                'max_wait_time': round(max_wait_time, 1),
                'avg_order_wait_time': round(avg_order_wait_time, 1),
                'max_order_wait_time': round(max_order_wait_time, 1),
                'avg_service_wait_time': round(avg_service_wait_time, 1),
                'max_service_wait_time': round(max_service_wait_time, 1),
                'active_channels': len(active_channels),
                'channels': active_channels
            }
        
        else:
            analytics = {
                'module': module_name,
                'error': 'Module not found'
            }
        
        return jsonify({'success': True, 'analytics': analytics})
    
    except Exception as e:
        logger.error(f"Error getting module analytics: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_fo otfall_report/<channel_id>')
def get_footfall_report(channel_id):
    """Get footfall report for a channel"""
    period = request.args.get('period', '7days')
    try:
        report_data = db_manager.get_footfall_report(channel_id, period)
        return jsonify(report_data)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_queue_report/<channel_id>')
def get_queue_report(channel_id):
    """Get queue analytics report"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    try:
        report_data = db_manager.get_queue_report(channel_id, start_date, end_date)
        return jsonify(report_data)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/get_alert_gifs')
def get_alert_gifs():
    """Get alert GIFs with optional filtering"""
    channel_id = request.args.get('channel_id')
    alert_type = request.args.get('alert_type')
    limit = int(request.args.get('limit', 20))
    days = request.args.get('days')
    days = int(days) if days else None
    
    try:
        alert_gifs = db_manager.get_alert_gifs(channel_id, alert_type, limit, days=days)
        return jsonify({
            'success': True,
            'alert_gifs': alert_gifs,
            'count': len(alert_gifs)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_alert_gif/<int:gif_id>', methods=['DELETE'])
@login_required
def delete_alert_gif(gif_id):
    """Delete an alert GIF"""
    try:
        with app.app_context():
            success = db_manager.delete_alert_gif(gif_id)
        if success:
            return jsonify({'success': True, 'message': 'Alert GIF deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Alert GIF not found'})
    except Exception as e:
        logger.error(f"Error deleting alert GIF {gif_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_old_alerts', methods=['POST'])
def clear_old_alerts():
    """Clear old queue monitor and bag detection alert GIFs"""
    data = request.json
    days = data.get('days', 7)
    alert_type = data.get('alert_type', 'all')  # 'queue_alert', 'bag_unattended', or 'all'
    
    try:
        with app.app_context():
            deleted_count = db_manager.cleanup_old_alert_gifs(max_age_days=days, alert_type=alert_type)
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Deleted {deleted_count} alerts older than {days} days'
        })
    except Exception as e:
        logger.error(f"Error clearing old alerts: {e}")
        return jsonify({'success': False, 'error': str(e)})

# Heatmap Routes
@app.route('/api/get_heatmap_snapshots')
def get_heatmap_snapshots():
    """Get heatmap snapshots with optional filtering"""
    channel_id = request.args.get('channel_id')
    limit = int(request.args.get('limit', 20))
    
    try:
        snapshots = db_manager.get_heatmap_snapshots(channel_id, limit)
        return jsonify({
            'success': True,
            'snapshots': snapshots,
            'count': len(snapshots)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_heatmap_snapshot/<int:snapshot_id>', methods=['DELETE'])
def delete_heatmap_snapshot(snapshot_id):
    """Delete a heatmap snapshot"""
    try:
        success = db_manager.delete_heatmap_snapshot(snapshot_id)
        if success:
            return jsonify({'success': True, 'message': 'Heatmap snapshot deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Snapshot not found'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/capture_heatmap_snapshot', methods=['POST'])
def capture_heatmap_snapshot():
    """Manually capture a heatmap snapshot"""
    data = request.json
    channel_id = data.get('channel_id')
    
    if not channel_id:
        return jsonify({'success': False, 'error': 'Channel ID is required'})
    
    try:
        # Check if Heatmap module is running for this channel
        if channel_id not in channel_modules or 'Heatmap' not in channel_modules[channel_id]:
            return jsonify({'success': False, 'error': 'Heatmap not running on this channel'})
        
        heatmap_module = channel_modules[channel_id]['Heatmap']
        
        # Get snapshot frame from the module
        snapshot_frame = heatmap_module.get_snapshot_frame()
        
        if snapshot_frame is None:
            return jsonify({'success': False, 'error': 'Failed to capture snapshot'})
        
        # Save the snapshot
        import cv2
        import os
        from datetime import datetime
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"heatmap_{channel_id}_{timestamp}.jpg"
        filepath = os.path.join('static/heatmaps', filename)
        
        # Ensure directory exists
        os.makedirs('static/heatmaps', exist_ok=True)
        
        # Save image
        cv2.imwrite(filepath, snapshot_frame)
        
        # Get file size
        file_size = os.path.getsize(filepath)
        
        # Save to database
        with app.app_context():
            snapshot_id = db_manager.save_heatmap_snapshot(
                channel_id=channel_id,
                filename=filename,
                filepath=filepath,
                file_size=file_size
            )
        
        return jsonify({
            'success': True,
            'message': 'Snapshot captured successfully',
            'snapshot_id': snapshot_id,
            'filename': filename
        })
        
    except Exception as e:
        logger.error(f"Error capturing heatmap snapshot: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_old_heatmap_snapshots', methods=['POST'])
def clear_old_heatmap_snapshots():
    """Clear old heatmap snapshots"""
    data = request.json
    days = data.get('days', 7)
    
    try:
        with app.app_context():
            deleted_count = db_manager.clear_old_heatmap_snapshots(days)
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Deleted {deleted_count} snapshots older than {days} days'
        })
    except Exception as e:
        logger.error(f"Error clearing old heatmap snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_heatmap_settings', methods=['POST'])
def set_heatmap_settings():
    """Set heatmap settings (decay rate, snapshot interval, intensity)"""
    data = request.json
    channel_id = data.get('channel_id')
    decay_rate = data.get('decay_rate')
    snapshot_interval = data.get('snapshot_interval')
    intensity = data.get('intensity')
    
    if not channel_id:
        return jsonify({'success': False, 'error': 'Channel ID is required'})
    
    try:
        if channel_id not in channel_modules or 'Heatmap' not in channel_modules[channel_id]:
            return jsonify({'success': False, 'error': 'Heatmap not running on this channel'})
        
        heatmap_module = channel_modules[channel_id]['Heatmap']
        
        # Update module settings
        if decay_rate is not None:
            heatmap_module.decay_rate = float(decay_rate)
        if snapshot_interval is not None:
            heatmap_module.snapshot_interval = int(snapshot_interval)
        if intensity is not None:
            heatmap_module.intensity = float(intensity)
        
        return jsonify({
            'success': True,
            'message': 'Heatmap settings updated successfully',
            'settings': {
                'decay_rate': heatmap_module.decay_rate,
                'snapshot_interval': heatmap_module.snapshot_interval,
                'intensity': heatmap_module.intensity
            }
        })
        
    except Exception as e:
        logger.error(f"Error setting heatmap settings: {e}")
        return jsonify({'success': False, 'error': str(e)})

# Configuration Management Routes
@app.route('/api/save_config', methods=['POST'])
def save_config():
    """Save channel configuration (ROI, counting line, etc.)"""
    try:
        data = request.json
        channel_id = data.get('channel_id')
        app_name = data.get('app_name')  # 'PeopleCounter' or 'QueueMonitor'
        config_type = data.get('config_type')  # 'roi' or 'counting_line'
        config_data = data.get('config_data')
        
        if not all([channel_id, app_name, config_type, config_data]):
            return jsonify({'success': False, 'error': 'Missing required fields'})
        
        logger.info(f"💾 Saving configuration: {channel_id} - {app_name} - {config_type}")
        logger.info(f"Configuration Data:\n{json.dumps(config_data, indent=2)}")
        
        db_manager.save_channel_config(channel_id, app_name, config_type, config_data)
        logger.info(f"✓ Configuration saved successfully to database")
        
        return jsonify({'success': True, 'message': 'Configuration saved successfully'})
    except Exception as e:
        logger.error(f"Error saving configuration: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_config/<channel_id>/<app_name>/<config_type>')
def get_config(channel_id, app_name, config_type):
    """Get channel configuration"""
    try:
        config_data = db_manager.get_channel_config(channel_id, app_name, config_type)
        if config_data:
            return jsonify({'success': True, 'config': config_data})
        else:
            return jsonify({'success': False, 'message': 'No configuration found'})
    except Exception as e:
        logger.error(f"Error retrieving configuration: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_all_configs/<channel_id>/<app_name>')
def get_all_configs(channel_id, app_name):
    """Get all configurations for a channel and app"""
    try:
        roi_config = db_manager.get_channel_config(channel_id, app_name, 'roi')
        line_config = db_manager.get_channel_config(channel_id, app_name, 'counting_line')
        
        return jsonify({
            'success': True,
            'configs': {
                'roi': roi_config,
                'counting_line': line_config
            }
        })
    except Exception as e:
        logger.error(f"Error retrieving configurations: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/update_bag_detection_config', methods=['POST'])
def update_bag_detection_config():
    """Update bag detection configuration"""
    try:
        data = request.json
        channel_id = data.get('channel_id')
        config = data.get('config', {})
        
        if not channel_id:
            return jsonify({'success': False, 'error': 'Channel ID required'})
        
        # Update the active module if it exists
        if channel_id in channel_modules and 'BagDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['BagDetection']
            module.update_config(config)
            return jsonify({'success': True, 'message': 'Configuration updated'})
        else:
            return jsonify({'success': False, 'error': 'Bag detection not running on this channel'})
            
    except Exception as e:
        logger.error(f"Error updating bag detection config: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_bag_detection_stats/<channel_id>')
def get_bag_detection_stats(channel_id):
    """Get bag detection statistics"""
    try:
        if channel_id in channel_modules and 'BagDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['BagDetection']
            stats = module.get_statistics()
            return jsonify({'success': True, 'stats': stats})
        else:
            return jsonify({'success': False, 'error': 'Bag detection not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Cash Detection Routes
@app.route('/static/cash_snapshots/<filename>')
def serve_cash_snapshot(filename):
    """Serve cash detection snapshot files"""
    return send_from_directory('static/cash_snapshots', filename)

@app.route('/api/get_cash_snapshots')
def get_cash_snapshots():
    """Get cash detection snapshots with optional filtering"""
    channel_id = request.args.get('channel_id')
    limit = int(request.args.get('limit', 50))
    
    try:
        snapshots = db_manager.get_cash_snapshots(channel_id, limit)
        return jsonify({
            'success': True,
            'snapshots': snapshots,
            'count': len(snapshots)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_cash_snapshot/<int:snapshot_id>', methods=['DELETE'])
def delete_cash_snapshot(snapshot_id):
    """Delete a cash detection snapshot"""
    try:
        success = db_manager.delete_cash_snapshot(snapshot_id)
        if success:
            return jsonify({'success': True, 'message': 'Cash snapshot deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Snapshot not found'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_old_cash_snapshots', methods=['POST'])
def clear_old_cash_snapshots():
    """Clear old cash detection snapshots"""
    data = request.json
    days = data.get('days', 7)
    
    try:
        with app.app_context():
            deleted_count = db_manager.clear_old_cash_snapshots(days)
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Deleted {deleted_count} snapshots older than {days} days'
        })
    except Exception as e:
        logger.error(f"Error clearing old cash snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/update_cash_detection_config', methods=['POST'])
def update_cash_detection_config():
    """Update cash detection configuration"""
    try:
        data = request.json
        channel_id = data.get('channel_id')
        config = data.get('config', {})
        
        if not channel_id:
            return jsonify({'success': False, 'error': 'Channel ID required'})
        
        # Update the active module if it exists
        if channel_id in channel_modules and 'CashDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['CashDetection']
            module.update_config(config)
            return jsonify({'success': True, 'message': 'Configuration updated'})
        else:
            return jsonify({'success': False, 'error': 'Cash detection not running on this channel'})
            
    except Exception as e:
        logger.error(f"Error updating cash detection config: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_cash_detection_stats/<channel_id>')
def get_cash_detection_stats(channel_id):
    """Get cash detection statistics"""
    try:
        if channel_id in channel_modules and 'CashDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['CashDetection']
            stats = module.get_statistics()
            return jsonify({'success': True, 'stats': stats})
        else:
            return jsonify({'success': False, 'error': 'Cash detection not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/export_cash_detection_excel')
def export_cash_detection_excel():
    """Export cash detection summary to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Get parameters
        channel_id = request.args.get('channel_id', None)
        days = int(request.args.get('days', 30))  # Default to last 30 days
        
        # Calculate date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        
        # Get all cash snapshots within date range
        with app.app_context():
            query = db_manager.CashSnapshot.query.filter(
                db_manager.CashSnapshot.created_at >= start_date,
                db_manager.CashSnapshot.created_at <= end_date
            )
            
            if channel_id:
                query = query.filter_by(channel_id=channel_id)
            
            snapshots = query.order_by(db_manager.CashSnapshot.created_at.desc()).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Detection Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="238636", end_color="238636", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title row
        ws.merge_cells('A1:H1')
        ws['A1'] = f'Cash Detection Summary Report ({start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")})'
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # Summary section
        row = 3
        ws[f'A{row}'] = 'Summary Statistics'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Calculate summary statistics
        total_detections = len(snapshots)
        total_channels = len(set(s.channel_id for s in snapshots))
        total_detection_count = sum(s.detection_count for s in snapshots)
        
        # Group by channel
        channel_stats = {}
        for snap in snapshots:
            ch_id = snap.channel_id
            if ch_id not in channel_stats:
                channel_stats[ch_id] = {'count': 0, 'total_detections': 0}
            channel_stats[ch_id]['count'] += 1
            channel_stats[ch_id]['total_detections'] += snap.detection_count
        
        # Summary data
        summary_data = [
            ['Total Detections', total_detections],
            ['Total Channels', total_channels],
            ['Total Detection Events', total_detection_count],
            ['Date Range', f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"],
            ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Channel-wise summary
        ws[f'A{row}'] = 'Channel-wise Summary'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Channel summary headers
        headers = ['Channel ID', 'Detection Count', 'Total Events']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row += 1
        
        # Channel summary data
        for ch_id, stats in sorted(channel_stats.items()):
            ws.cell(row, 1, ch_id).border = border
            ws.cell(row, 2, stats['count']).border = border
            ws.cell(row, 3, stats['total_detections']).border = border
            row += 1
        
        row += 2
        
        # Detailed data section
        ws[f'A{row}'] = 'Detailed Detection Records'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers for detailed data
        detail_headers = ['ID', 'Channel ID', 'Timestamp', 'Detection Count', 'Alert Message', 'File Name', 'File Size (KB)', 'Snapshot URL']
        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row += 1
        
        # Detailed data rows
        for snap in snapshots:
            file_size_kb = round(snap.file_size / 1024, 2) if snap.file_size else 0
            timestamp = snap.created_at.strftime('%Y-%m-%d %H:%M:%S') if snap.created_at else 'N/A'
            
            data_row = [
                snap.id,
                snap.channel_id,
                timestamp,
                snap.detection_count,
                snap.alert_message or 'N/A',
                snap.snapshot_filename,
                file_size_kb,
                f'/static/cash_snapshots/{snap.snapshot_filename}'
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row, col, value)
                cell.border = border
                if col == 3:  # Timestamp column
                    cell.alignment = Alignment(horizontal='left')
            
            row += 1
        
        # Auto-adjust column widths
        column_widths = {
            'A': 10,  # ID
            'B': 15,  # Channel ID
            'C': 20,  # Timestamp
            'D': 15,  # Detection Count
            'E': 30,  # Alert Message
            'F': 30,  # File Name
            'G': 15,  # File Size
            'H': 40   # Snapshot URL
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f'cash_detection_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        if channel_id:
            filename = f'cash_detection_summary_{channel_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except ImportError:
        logger.error("openpyxl not installed. Please install it: pip install openpyxl")
        return jsonify({'success': False, 'error': 'Excel export requires openpyxl. Please install it: pip install openpyxl'}), 500
    except Exception as e:
        logger.error(f"Error exporting cash detection Excel: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export_unauthorized_entry_excel')
@login_required
def export_unauthorized_entry_excel():
    """Export unauthorized entry summary to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        import json
        
        # Get parameters
        channel_id = request.args.get('channel_id', None)
        days = int(request.args.get('days', 30))  # Default to last 30 days
        
        # Calculate date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        
        # Get all unauthorized entry alerts within date range
        with app.app_context():
            query = db_manager.AlertGif.query.filter(
                db_manager.AlertGif.alert_type == 'unauthorized_entry_alert',
                db_manager.AlertGif.created_at >= start_date,
                db_manager.AlertGif.created_at <= end_date
            )
            
            if channel_id:
                query = query.filter_by(channel_id=channel_id)
            
            alerts = query.order_by(db_manager.AlertGif.created_at.desc()).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Unauthorized Entry Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")  # Red for security alerts
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title row
        ws.merge_cells('A1:H1')
        ws['A1'] = f'Unauthorized Entry Alert Summary Report ({start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")})'
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # Summary section
        row = 3
        ws[f'A{row}'] = 'Summary Statistics'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Calculate summary statistics
        total_alerts = len(alerts)
        total_channels = len(set(a.channel_id for a in alerts))
        
        # Group by channel
        channel_stats = {}
        for alert in alerts:
            ch_id = alert.channel_id
            if ch_id not in channel_stats:
                channel_stats[ch_id] = {'count': 0}
            channel_stats[ch_id]['count'] += 1
        
        # Summary data
        summary_data = [
            ['Total Alerts', total_alerts],
            ['Total Channels', total_channels],
            ['Date Range', f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"],
            ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 2
        
        # Channel-wise summary
        ws[f'A{row}'] = 'Channel-wise Summary'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Channel summary headers
        headers = ['Channel ID', 'Alert Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row += 1
        
        # Channel summary data
        for ch_id, stats in sorted(channel_stats.items()):
            ws.cell(row, 1, ch_id).border = border
            ws.cell(row, 2, stats['count']).border = border
            row += 1
        
        row += 2
        
        # Detailed data section
        ws[f'A{row}'] = 'Detailed Alert Records'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers for detailed data
        detail_headers = ['ID', 'Channel ID', 'Timestamp', 'Alert Message', 'File Name', 'File Size (KB)', 'Duration (s)', 'GIF URL']
        for col, header in enumerate(detail_headers, 1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row += 1
        
        # Detailed data rows
        for alert in alerts:
            file_size_kb = round(alert.file_size / 1024, 2) if alert.file_size else 0
            timestamp = alert.created_at.strftime('%Y-%m-%d %H:%M:%S') if alert.created_at else 'N/A'
            duration = alert.duration_seconds if alert.duration_seconds else 'N/A'
            
            # Parse alert_data if available
            alert_message = alert.alert_message or 'N/A'
            if alert.alert_data:
                try:
                    alert_data = json.loads(alert.alert_data) if isinstance(alert.alert_data, str) else alert.alert_data
                    if isinstance(alert_data, dict):
                        # Extract additional info from alert_data
                        if 'message' in alert_data:
                            alert_message = alert_data['message']
                        elif 'detection_count' in alert_data:
                            alert_message = f"Unauthorized entry detected: {alert_data['detection_count']} person(s)"
                except:
                    pass
            
            data_row = [
                alert.id,
                alert.channel_id,
                timestamp,
                alert_message,
                alert.gif_filename,
                file_size_kb,
                duration,
                f'/static/alerts/{alert.gif_filename}'
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row, col, value)
                cell.border = border
                if col == 3:  # Timestamp column
                    cell.alignment = Alignment(horizontal='left')
            
            row += 1
        
        # Auto-adjust column widths
        column_widths = {
            'A': 10,  # ID
            'B': 15,  # Channel ID
            'C': 20,  # Timestamp
            'D': 40,  # Alert Message
            'E': 30,  # File Name
            'F': 15,  # File Size
            'G': 15,  # Duration
            'H': 40   # GIF URL
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f'unauthorized_entry_summary_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        if channel_id:
            filename = f'unauthorized_entry_summary_{channel_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except ImportError:
        logger.error("openpyxl not installed. Please install it: pip install openpyxl")
        return jsonify({'success': False, 'error': 'Excel export requires openpyxl. Please install it: pip install openpyxl'}), 500
    except Exception as e:
        logger.error(f"Error exporting unauthorized entry Excel: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

# Fall Detection routes
@app.route('/static/fall_snapshots/<filename>')
def fall_snapshot(filename):
    """Serve fall detection snapshot images"""
    return send_from_directory('static/fall_snapshots', filename)

@app.route('/api/get_fall_snapshots')
def get_fall_snapshots():
    """Get fall detection snapshots"""
    try:
        channel_id = request.args.get('channel_id')
        limit = int(request.args.get('limit', 50))
        
        snapshots = db_manager.get_fall_snapshots(channel_id=channel_id, limit=limit)
        
        return jsonify({
            'success': True,
            'snapshots': snapshots
        })
    except Exception as e:
        logger.error(f"Error getting fall snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_fall_snapshot/<int:snapshot_id>', methods=['DELETE'])
def delete_fall_snapshot(snapshot_id):
    """Delete a fall detection snapshot"""
    try:
        success = db_manager.delete_fall_snapshot(snapshot_id)
        if success:
            return jsonify({'success': True, 'message': 'Snapshot deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Snapshot not found'})
    except Exception as e:
        logger.error(f"Error deleting fall snapshot: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_old_fall_snapshots', methods=['POST'])
def clear_old_fall_snapshots():
    """Clear old fall detection snapshots"""
    try:
        data = request.json or {}
        days = int(data.get('days', 7))
        
        deleted_count = db_manager.clear_old_fall_snapshots(days)
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Deleted {deleted_count} snapshots older than {days} days'
        })
    except Exception as e:
        logger.error(f"Error clearing old fall snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/update_fall_detection_config', methods=['POST'])
def update_fall_detection_config():
    """Update fall detection configuration"""
    try:
        data = request.json
        channel_id = data.get('channel_id')
        config = data.get('config', {})
        
        if not channel_id:
            return jsonify({'success': False, 'error': 'Channel ID required'})
        
        # Update the active module if it exists
        if channel_id in channel_modules and 'FallDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['FallDetection']
            module.update_config(config)
            return jsonify({'success': True, 'message': 'Configuration updated'})
        else:
            return jsonify({'success': False, 'error': 'Fall detection not running on this channel'})
            
    except Exception as e:
        logger.error(f"Error updating fall detection config: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_fall_detection_stats/<channel_id>')
def get_fall_detection_stats(channel_id):
    """Get fall detection statistics"""
    try:
        if channel_id in channel_modules and 'FallDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['FallDetection']
            stats = module.get_statistics()
            return jsonify({'success': True, 'stats': stats})
        else:
            return jsonify({'success': False, 'error': 'Fall detection not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# ============= Mopping Detection Endpoints =============
@app.route('/static/mopping_snapshots/<filename>')
def serve_mopping_snapshot(filename):
    """Serve mopping snapshot images"""
    return send_from_directory('static/mopping_snapshots', filename)

@app.route('/api/get_mopping_snapshots')
def get_mopping_snapshots():
    """Get mopping detection snapshots from database"""
    try:
        channel_id = request.args.get('channel_id')
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        
        snapshots = db_manager.get_mopping_snapshots(channel_id=channel_id, limit=limit, offset=offset)
        return jsonify({'success': True, 'snapshots': snapshots})
    except Exception as e:
        logger.error(f"Error getting mopping snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_mopping_snapshot/<int:snapshot_id>', methods=['DELETE'])
def delete_mopping_snapshot(snapshot_id):
    """Delete a mopping snapshot"""
    try:
        logger.info(f"Attempting to delete mopping snapshot ID: {snapshot_id}")
        success = db_manager.delete_mopping_snapshot(snapshot_id)
        if success:
            logger.info(f"Successfully deleted mopping snapshot ID: {snapshot_id}")
            return jsonify({'success': True, 'message': 'Snapshot deleted'})
        else:
            logger.warning(f"Mopping snapshot ID {snapshot_id} not found in database")
            return jsonify({'success': False, 'error': 'Snapshot not found'})
    except Exception as e:
        logger.error(f"Error deleting mopping snapshot {snapshot_id}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_old_mopping_snapshots', methods=['POST'])
def clear_old_mopping_snapshots():
    """Clear old mopping snapshots older than specified days"""
    try:
        data = request.json or {}
        days = data.get('days', 7)
        deleted_count = db_manager.clear_old_mopping_snapshots(days)
        return jsonify({
            'success': True, 
            'message': f'Cleared {deleted_count} snapshots older than {days} days',
            'deleted_count': deleted_count
        })
    except Exception as e:
        logger.error(f"Error clearing old mopping snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/update_mopping_detection_config', methods=['POST'])
def update_mopping_detection_config():
    """Update mopping detection configuration"""
    try:
        data = request.json
        channel_id = data.get('channel_id')
        
        if channel_id in channel_modules and 'MoppingDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['MoppingDetection']
            module.update_config(data)
            return jsonify({'success': True, 'message': 'Configuration updated'})
        else:
            return jsonify({'success': False, 'error': 'Mopping detection not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_mopping_detection_stats/<channel_id>')
def get_mopping_detection_stats(channel_id):
    """Get mopping detection statistics"""
    try:
        if channel_id in channel_modules and 'MoppingDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['MoppingDetection']
            stats = module.get_statistics()
            return jsonify({'success': True, 'stats': stats})
        else:
            return jsonify({'success': False, 'error': 'Mopping detection not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# ============= Smoking Detection Routes =============
@app.route('/static/smoking_snapshots/<filename>')
def serve_smoking_snapshot(filename):
    """Serve smoking detection snapshot images"""
    return send_from_directory('static/smoking_snapshots', filename)

@app.route('/api/get_smoking_snapshots')
def get_smoking_snapshots():
    """Get all smoking detection snapshots"""
    try:
        channel_id = request.args.get('channel_id')
        limit = int(request.args.get('limit', 50))
        
        snapshots = db_manager.get_smoking_snapshots(channel_id=channel_id, limit=limit)
        return jsonify({'success': True, 'snapshots': snapshots})
    except Exception as e:
        logger.error(f"Error getting smoking snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_smoking_snapshot/<int:snapshot_id>', methods=['DELETE'])
@login_required
def delete_smoking_snapshot(snapshot_id):
    """Delete a smoking detection snapshot"""
    try:
        with app.app_context():
            success = db_manager.delete_smoking_snapshot(snapshot_id)
            if success:
                return jsonify({'success': True, 'message': 'Snapshot deleted successfully'})
            else:
                return jsonify({'success': False, 'error': 'Snapshot not found'})
    except Exception as e:
        logger.error(f"Error deleting smoking snapshot: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_old_smoking_snapshots', methods=['POST'])
@login_required
def clear_old_smoking_snapshots():
    """Clear old smoking detection snapshots"""
    try:
        data = request.json or {}
        days = int(data.get('days', 7))
        
        with app.app_context():
            deleted_count = db_manager.clear_old_smoking_snapshots(days)
        
        return jsonify({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Deleted {deleted_count} snapshots older than {days} days'
        })
    except Exception as e:
        logger.error(f"Error clearing old smoking snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/update_smoking_detection_config', methods=['POST'])
def update_smoking_detection_config():
    """Update smoking detection configuration"""
    try:
        data = request.json
        channel_id = data.get('channel_id')
        config = data.get('config', {})
        
        # Update config
        return jsonify({'success': True, 'message': 'Configuration updated'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_smoking_detection_stats/<channel_id>')
def get_smoking_detection_stats(channel_id):
    """Get smoking detection statistics"""
    try:
        if channel_id in channel_modules and 'SmokingDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['SmokingDetection']
            stats = module.get_statistics()
            return jsonify({'success': True, 'stats': stats})
        else:
            return jsonify({'success': False, 'error': 'Smoking detection not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== Phone Usage Detection Routes ====================

@app.route('/static/phone_snapshots/<filename>')
def serve_phone_snapshot(filename):
    """Serve phone usage detection snapshot images"""
    return send_from_directory('static/phone_snapshots', filename)

@app.route('/api/get_phone_snapshots')
def get_phone_snapshots():
    """Get all phone usage detection snapshots"""
    try:
        channel_id = request.args.get('channel_id')
        limit = int(request.args.get('limit', 50))
        
        snapshots = db_manager.get_phone_snapshots(channel_id=channel_id, limit=limit)
        return jsonify({'success': True, 'snapshots': snapshots})
    except Exception as e:
        logger.error(f"Error getting phone snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_phone_snapshot/<int:snapshot_id>', methods=['DELETE'])
def delete_phone_snapshot(snapshot_id):
    """Delete a phone usage detection snapshot"""
    try:
        success = db_manager.delete_phone_snapshot(snapshot_id)
        if success:
            return jsonify({'success': True, 'message': 'Snapshot deleted'})
        else:
            return jsonify({'success': False, 'error': 'Snapshot not found'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_old_phone_snapshots', methods=['POST'])
def clear_old_phone_snapshots():
    """Clear old phone usage detection snapshots"""
    try:
        days = request.json.get('days', 7)
        count = db_manager.clear_old_phone_snapshots(days=days)
        return jsonify({'success': True, 'message': f'Cleared {count} snapshots older than {days} days'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_phone_detection_stats/<channel_id>')
def get_phone_detection_stats(channel_id):
    """Get phone usage detection statistics"""
    try:
        if channel_id in channel_modules and 'PhoneUsageDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['PhoneUsageDetection']
            stats = module.get_statistics()
            return jsonify({'success': True, 'stats': stats})
        else:
            return jsonify({'success': False, 'error': 'Phone detection not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Restricted Area Monitor API Routes
@app.route('/api/get_restricted_area_snapshots')
def get_restricted_area_snapshots():
    """Get restricted area violation snapshots"""
    try:
        channel_id = request.args.get('channel_id')
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        
        snapshots = db_manager.get_restricted_area_snapshots(channel_id, limit, offset)
        return jsonify({'success': True, 'snapshots': snapshots})
    except Exception as e:
        logger.error(f"Error getting restricted area snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_restricted_area_snapshot/<int:snapshot_id>', methods=['DELETE'])
def delete_restricted_area_snapshot(snapshot_id):
    """Delete a restricted area snapshot"""
    try:
        logger.info(f"Attempting to delete restricted area snapshot ID: {snapshot_id}")
        success = db_manager.delete_restricted_area_snapshot(snapshot_id)
        if success:
            logger.info(f"Successfully deleted restricted area snapshot ID: {snapshot_id}")
            return jsonify({'success': True, 'message': 'Snapshot deleted'})
        else:
            logger.warning(f"Restricted area snapshot ID {snapshot_id} not found in database")
            return jsonify({'success': False, 'error': 'Snapshot not found'})
    except Exception as e:
        logger.error(f"Error deleting restricted area snapshot {snapshot_id}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_old_restricted_area_snapshots', methods=['POST'])
def clear_old_restricted_area_snapshots():
    """Clear old restricted area snapshots"""
    try:
        data = request.json or {}
        days = data.get('days', 7)
        deleted_count = db_manager.clear_old_restricted_area_snapshots(days)
        return jsonify({
            'success': True,
            'message': f'Cleared {deleted_count} snapshots older than {days} days',
            'deleted_count': deleted_count
        })
    except Exception as e:
        logger.error(f"Error clearing old restricted area snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/static/restricted_area_snapshots/<filename>')
def serve_restricted_area_snapshot(filename):
    """Serve restricted area snapshot file"""
    return send_from_directory('static/restricted_area_snapshots', filename)

@app.route('/api/get_restricted_area_stats/<channel_id>')
def get_restricted_area_stats(channel_id):
    """Get restricted area monitoring statistics for a channel"""
    try:
        if channel_id in channel_modules and 'RestrictedAreaMonitor' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['RestrictedAreaMonitor']
            stats = module.get_statistics()
            return jsonify({'success': True, 'stats': stats})
        else:
            return jsonify({'success': False, 'error': 'Restricted area monitor not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_restricted_area_roi/<channel_id>', methods=['POST'])
def set_restricted_area_roi(channel_id):
    """Set ROI points for restricted area monitor"""
    try:
        data = request.json
        roi_points = data.get('roi_points', [])
        
        if channel_id in channel_modules and 'RestrictedAreaMonitor' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['RestrictedAreaMonitor']
            module.set_roi_points(roi_points)
            
            # Save to database using the generic config method
            db_manager.save_channel_config(channel_id, 'RestrictedAreaMonitor', 'roi', roi_points)
            logger.info(f"✓ ROI points saved to database for RestrictedAreaMonitor on channel {channel_id}")
            
            return jsonify({'success': True, 'message': 'ROI points saved'})
        else:
            return jsonify({'success': False, 'error': 'Restricted area monitor not running'})
    except Exception as e:
        logger.error(f"Error setting restricted area ROI: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_restricted_area_roi/<channel_id>')
def get_restricted_area_roi(channel_id):
    """Get ROI points for restricted area monitor"""
    try:
        if channel_id in channel_modules and 'RestrictedAreaMonitor' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['RestrictedAreaMonitor']
            roi_points = module.get_roi_points()
            return jsonify({'success': True, 'roi_points': roi_points})
        else:
            return jsonify({'success': False, 'error': 'Restricted area monitor not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/model_stats')
@login_required
def get_model_stats_api():
    """Get global model manager statistics"""
    try:
        stats = get_model_stats()
        return jsonify({'success': True, 'stats': stats})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/cleanup_models', methods=['POST'])
@login_required
def cleanup_models_api():
    """Clean up unused models"""
    try:
        data = request.json
        max_age = data.get('max_age_seconds', 3600)  # Default 1 hour
        
        cleaned_count = cleanup_models(max_age)
        return jsonify({
            'success': True, 
            'message': f'Cleaned up {cleaned_count} unused models',
            'cleaned_count': cleaned_count
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_grooming_detection_stats/<channel_id>')
def get_grooming_detection_stats(channel_id):
    """Get grooming detection statistics"""
    try:
        if channel_id in channel_modules and 'GroomingDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['GroomingDetection']
            stats = module.get_statistics()
            return jsonify({'success': True, 'stats': stats})
        else:
            return jsonify({'success': False, 'error': 'Grooming detection not running'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_crowd_roi/<channel_id>', methods=['POST'])
def set_crowd_roi(channel_id):
    """Set ROI points for crowd detection"""
    try:
        data = request.json
        roi_points = data.get('roi_points', [])
        
        if channel_id in channel_modules and 'CrowdDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['CrowdDetection']
            # Convert list to dict format if needed
            if isinstance(roi_points, list):
                module.set_roi({'main': roi_points})
            elif isinstance(roi_points, dict):
                module.set_roi(roi_points)
            else:
                return jsonify({'success': False, 'error': 'Invalid ROI format'})
            
            logger.info(f"✓ ROI points saved for CrowdDetection on channel {channel_id}")
            
            return jsonify({'success': True, 'message': 'ROI points saved'})
        else:
            return jsonify({'success': False, 'error': 'Crowd detection not running'})
    except Exception as e:
        logger.error(f"Error setting crowd ROI: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_crowd_roi/<channel_id>')
def get_crowd_roi(channel_id):
    """Get ROI points for crowd detection"""
    try:
        if channel_id in channel_modules and 'CrowdDetection' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['CrowdDetection']
            roi_config = module.get_roi()
            # Return main ROI points as a list for consistency
            roi_points = roi_config.get('main', [])
            return jsonify({'success': True, 'roi_points': roi_points})
        else:
            return jsonify({'success': False, 'error': 'Crowd detection not running'})
    except Exception as e:
        logger.error(f"Error getting crowd ROI: {e}")
        return jsonify({'success': False, 'error': str(e)})

# Socket.IO events
# Track which clients are subscribed to which video streams
active_stream_subscriptions = {}  # {session_id: {channel_id: True}}
stream_broadcast_threads = {}  # {channel_id: thread}
stream_broadcast_stop_flags = {}  # {channel_id: threading.Event}
stream_subscriber_counts = {}  # {stream_key: int}
stream_broadcast_lock = threading.Lock()

@socketio.on('connect')
def handle_connect():
    logger.info('Client connected')
    emit('status', {'message': 'Connected to Sakshi.AI'})
    # Initialize subscription tracking for this client
    active_stream_subscriptions[request.sid] = {}

@socketio.on('disconnect')
def handle_disconnect():
    logger.info('Client disconnected')
    # Decrement subscriber counts for any streams this client was subscribed to
    try:
        client_subs = active_stream_subscriptions.get(request.sid, {})
        for stream_key in list(client_subs.keys()):
            with stream_broadcast_lock:
                stream_subscriber_counts[stream_key] = max(0, stream_subscriber_counts.get(stream_key, 0) - 1)
    except Exception:
        pass
    # Clean up subscriptions for this client
    if request.sid in active_stream_subscriptions:
        del active_stream_subscriptions[request.sid]

@socketio.on('subscribe_stream')
def handle_subscribe_stream(data):
    """Client subscribes to a video stream"""
    try:
        app_name = data.get('app_name')
        channel_id = data.get('channel_id')
        
        if not app_name or not channel_id:
            emit('stream_error', {'error': 'Missing app_name or channel_id'})
            return

        # Acknowledge subscription immediately so the UI doesn't show "No Signal"
        # while we (re)start processors and wait for first frames.
        emit('stream_subscribed', {'app_name': app_name, 'channel_id': channel_id})
        
        # Track this subscription
        if request.sid not in active_stream_subscriptions:
            active_stream_subscriptions[request.sid] = {}
        stream_key = f"{app_name}:{channel_id}"
        # Avoid double-counting if the UI emits duplicate subscribe events
        if not active_stream_subscriptions[request.sid].get(stream_key):
            active_stream_subscriptions[request.sid][stream_key] = True
            with stream_broadcast_lock:
                stream_subscriber_counts[stream_key] = stream_subscriber_counts.get(stream_key, 0) + 1
        
        logger.info(f"Client {request.sid} subscribed to {app_name}/{channel_id}")
        
        # Check if channel processor exists and is running
        if channel_id not in shared_video_processors:
            # Channel is configured but processor not running - try to restart it
            logger.info(f"Channel {channel_id} not in shared_video_processors, attempting to restart...")
            
            # Get channel configuration - channels.json is source of truth
            rtsp_url = None
            
            # FIRST: Always check channels.json (source of truth) and update database
            try:
                config_path = Path('config/channels.json')
                if config_path.exists():
                    with open(config_path, 'r') as f:
                        config = json.load(f)
                        for ch in config.get('channels', []):
                            if ch.get('channel_id') == channel_id:
                                rtsp_url = ch.get('rtsp_url')
                                if rtsp_url:
                                    logger.info(f"Found RTSP URL for {channel_id} in channels.json: {rtsp_url}")
                                    # Update database with correct URL from channels.json
                                    try:
                                        channel_name = ch.get('channel_name', channel_id)
                                        with app.app_context():
                                            db_manager.save_rtsp_channel(
                                                channel_id, 
                                                channel_name, 
                                                rtsp_url,
                                                description=f"Auto-updated from channels.json"
                                            )
                                            logger.info(f"💾 Updated database with correct RTSP URL for {channel_id}")
                                    except Exception as db_error:
                                        logger.warning(f"Could not update database URL for {channel_id}: {db_error}")
                                break
            except Exception as e:
                logger.error(f"Error reading channels.json: {e}")
            
            # Fallback to database only if not found in channels.json
            if not rtsp_url:
                try:
                    with app.app_context():
                        rtsp_channel = db_manager.get_rtsp_channel(channel_id)
                        if rtsp_channel:
                            rtsp_url = rtsp_channel.get('rtsp_url')
                            if rtsp_url:
                                logger.info(f"Found RTSP URL for {channel_id} in database (fallback)")
                except Exception as e:
                    logger.debug(f"Could not get RTSP URL from database: {e}")
            
            # If still not found, try app_configs
            if not rtsp_url and channel_id in channel_modules:
                for module_type in channel_modules[channel_id].keys():
                    if module_type in app_configs and channel_id in app_configs[module_type]['channels']:
                        rtsp_url = app_configs[module_type]['channels'][channel_id].get('video_source')
                        if rtsp_url:
                            logger.info(f"Found RTSP URL for {channel_id} in app_configs (fallback)")
                            break
            
            if rtsp_url:
                # Attempt to create and start processor
                try:
                    logger.info(f"Attempting to start processor for {channel_id} with RTSP: {rtsp_url}")
                    processor = SharedMultiModuleVideoProcessor(
                        video_source=rtsp_url,
                        channel_id=channel_id,
                        fps_limit=30
                    )
                    shared_video_processors[channel_id] = processor
                    
                    # Add existing modules back to processor
                    if channel_id in channel_modules and channel_modules[channel_id]:
                        for module_type, module in channel_modules[channel_id].items():
                            processor.add_module(module_type, module)
                    else:
                        # If we don't have cached module instances for this channel yet,
                        # create at least the requested module so the stream can render.
                        # This happens when a channel was never started via channels.json
                        # load but a user opens it directly from the dashboard.
                        try:
                            if channel_id not in channel_modules:
                                channel_modules[channel_id] = {}
                            if app_name == 'MaterialTheftMonitor':
                                module = MaterialTheftMonitor(channel_id, socketio, db_manager, app)
                                processor.add_module('MaterialTheftMonitor', module)
                                channel_modules[channel_id]['MaterialTheftMonitor'] = module
                                logger.info(f"✅ Created MaterialTheftMonitor instance for {channel_id} on-demand")
                            # Add more module types here as needed
                        except Exception as e:
                            logger.error(f"Failed to create module {app_name} for {channel_id}: {e}")
                    
                    # Start the processor
                    if processor.start():
                        logger.info(f"✓ Successfully restarted channel {channel_id}")
                    else:
                        logger.warning(f"⚠ Failed to start channel {channel_id} - RTSP connection may be unavailable")
                        emit('stream_error', {
                            'app_name': app_name,
                            'channel_id': channel_id,
                            'error': f'Channel {channel_id} is configured but RTSP connection failed. Please check camera connectivity.'
                        })
                except Exception as e:
                    logger.error(f"Error restarting channel {channel_id}: {e}")
                    emit('stream_error', {
                        'app_name': app_name,
                        'channel_id': channel_id,
                        'error': f'Failed to restart channel {channel_id}: {str(e)}'
                    })
            else:
                logger.warning(f"Could not find RTSP URL for channel {channel_id}")
                emit('stream_error', {
                    'app_name': app_name,
                    'channel_id': channel_id,
                    'error': f'Channel {channel_id} is configured but RTSP URL not found. Cannot start video stream.'
                })
        elif not shared_video_processors[channel_id].is_running:
            # Processor exists but is_running flag is False - check if it's actually working
            processor = shared_video_processors[channel_id]
            
            # Check multiple indicators that processor is actually running
            thread_alive = hasattr(processor, 'processing_thread') and processor.processing_thread and processor.processing_thread.is_alive()
            has_recent_frame = False
            if hasattr(processor, 'latest_raw_frame') and processor.latest_raw_frame is not None:
                has_recent_frame = True
            elif hasattr(processor, 'latest_annotated_frame') and processor.latest_annotated_frame is not None:
                has_recent_frame = True
            
            if thread_alive or has_recent_frame:
                # Processor is actually working, just flag might be wrong or timing issue
                logger.info(f"Channel {channel_id} processor appears to be working (thread_alive={thread_alive}, has_frame={has_recent_frame}), allowing connection")
                # Continue with broadcast setup below - processor is functional
            else:
                # Processor truly not running - try to restart it with correct URL from channels.json
                logger.warning(f"Channel {channel_id} processor exists but is not running (thread not alive, no recent frames). Attempting restart...")
                
                # Get correct RTSP URL from channels.json (source of truth)
                rtsp_url = None
                try:
                    config_path = Path('config/channels.json')
                    if config_path.exists():
                        with open(config_path, 'r') as f:
                            config = json.load(f)
                            for ch in config.get('channels', []):
                                if ch.get('channel_id') == channel_id:
                                    rtsp_url = ch.get('rtsp_url')
                                    if rtsp_url:
                                        logger.info(f"Found RTSP URL for {channel_id} in channels.json: {rtsp_url}")
                                    break
                except Exception as e:
                    logger.error(f"Error reading channels.json: {e}")
                
                if rtsp_url:
                    try:
                        # Stop and remove old processor (if it exists)
                        if channel_id in shared_video_processors:
                            old_processor = shared_video_processors[channel_id]
                            try:
                                old_processor.stop()
                            except:
                                pass
                            del shared_video_processors[channel_id]
                        
                        # Create new processor with correct URL
                        logger.info(f"Restarting processor for {channel_id} with RTSP URL from channels.json")
                        processor = SharedMultiModuleVideoProcessor(
                            video_source=rtsp_url,
                            channel_id=channel_id,
                            fps_limit=30
                        )
                        shared_video_processors[channel_id] = processor
                        
                        # Re-add modules
                        if channel_id in channel_modules and channel_modules[channel_id]:
                            for module_type, module in channel_modules[channel_id].items():
                                processor.add_module(module_type, module)
                        
                        # Start processor
                        if processor.start():
                            logger.info(f"✓ Successfully restarted channel {channel_id} with URL from channels.json")
                            # Continue with broadcast setup below
                        else:
                            logger.warning(f"⚠ Failed to restart channel {channel_id} - RTSP connection may be unavailable")
                            emit('stream_error', {
                                'app_name': app_name,
                                'channel_id': channel_id,
                                'error': f'Channel {channel_id} restart failed. RTSP connection may be unavailable.'
                            })
                            return
                    except Exception as e:
                        logger.error(f"Error restarting channel {channel_id}: {e}", exc_info=True)
                        emit('stream_error', {
                            'app_name': app_name,
                            'channel_id': channel_id,
                            'error': f'Failed to restart channel {channel_id}: {str(e)}'
                        })
                        return
                else:
                    logger.warning(f"Could not find RTSP URL for {channel_id} in channels.json")
                    emit('stream_error', {
                        'app_name': app_name,
                        'channel_id': channel_id,
                        'error': f'Channel {channel_id} is not running and RTSP URL not found in channels.json.'
                    })
                    return
        
        # Start broadcast thread for this channel if not already running
        # Start broadcast thread for this channel if not already running.
        # Protected by a lock to avoid duplicate thread creation when the UI emits
        # multiple subscribe events quickly (which causes "No Signal" flapping).
        with stream_broadcast_lock:
            existing = stream_broadcast_threads.get(stream_key)
            if existing is None or not existing.is_alive():
                stop_flag = threading.Event()
                stream_broadcast_stop_flags[stream_key] = stop_flag
                thread = threading.Thread(
                    target=broadcast_video_frames,
                    args=(app_name, channel_id, stop_flag),
                    daemon=True
                )
                thread.start()
                stream_broadcast_threads[stream_key] = thread
                logger.info(f"Started broadcast thread for {stream_key}")
        
        # Note: stream_subscribed is emitted above immediately.
        
    except Exception as e:
        logger.error(f"Error subscribing to stream: {e}")
        emit('stream_error', {'error': str(e)})

@socketio.on('unsubscribe_stream')
def handle_unsubscribe_stream(data):
    """Client unsubscribes from a video stream"""
    try:
        app_name = data.get('app_name')
        channel_id = data.get('channel_id')
        
        if not app_name or not channel_id:
            return
        
        # Remove subscription
        stream_key = f"{app_name}:{channel_id}"
        removed = False
        if request.sid in active_stream_subscriptions and active_stream_subscriptions[request.sid].pop(stream_key, None):
            removed = True
        if removed:
            with stream_broadcast_lock:
                stream_subscriber_counts[stream_key] = max(0, stream_subscriber_counts.get(stream_key, 0) - 1)
        
        logger.info(f"Client {request.sid} unsubscribed from {app_name}/{channel_id}")
        # Do NOT stop the broadcast thread immediately: the UI frequently unsubscribes/re-subscribes
        # during tab switches/pagination, which causes "No Signal" flapping.
        # The broadcast loop will self-terminate after a grace period with zero subscribers.
        with stream_broadcast_lock:
            subs = stream_subscriber_counts.get(stream_key, 0)
        if subs <= 0:
            logger.info(f"{stream_key} now has 0 subscribers (will stop after grace period)")
        
    except Exception as e:
        logger.error(f"Error unsubscribing from stream: {e}")

def broadcast_video_frames(app_name, channel_id, stop_flag):
    """Broadcast video frames to subscribed clients via Socket.IO"""
    import base64
    stream_key = f"{app_name}:{channel_id}"
    frame_interval = 1.0 / 15  # 15 FPS for web streaming
    last_frame_time = 0
    no_frame_start_time = None
    error_emitted = False
    consecutive_none_frames = 0
    max_consecutive_none = 300  # Allow up to 300 consecutive None frames before giving up
    
    logger.info(f"Broadcasting frames for {stream_key}")
    last_subscriber_time = time.time()
    no_subscriber_grace_seconds = 15.0
    
    while not stop_flag.is_set():
        try:
            # Self-stop after a grace period with no subscribers.
            with stream_broadcast_lock:
                subs = stream_subscriber_counts.get(stream_key, 0)
            if subs > 0:
                last_subscriber_time = time.time()
            elif (time.time() - last_subscriber_time) > no_subscriber_grace_seconds:
                logger.info(f"Stopping broadcast thread for {stream_key} (0 subscribers for {no_subscriber_grace_seconds:.0f}s)")
                break

            current_time = time.time()
            
            # Rate limiting
            if current_time - last_frame_time < frame_interval:
                time.sleep(0.01)
                continue
            
            # Get frame from video processor
            if channel_id not in shared_video_processors:
                if no_frame_start_time is None:
                    no_frame_start_time = current_time
                elif current_time - no_frame_start_time > 3.0 and not error_emitted:
                    # No processor after 3 seconds - emit error
                    socketio.emit('stream_error', {
                        'app_name': app_name,
                        'channel_id': channel_id,
                        'error': f'Channel {channel_id} is not running. RTSP connection may have failed. Please check camera connectivity.'
                    }, room=None)
                    error_emitted = True
                time.sleep(0.5)
                continue
            
            processor = shared_video_processors[channel_id]
            
            # Check if processor is actually working (not just is_running flag)
            # A processor might be processing frames even if is_running is False
            is_actually_working = False
            if processor.is_running:
                is_actually_working = True
            else:
                # Check if thread is alive or has recent frames
                thread_alive = hasattr(processor, 'processing_thread') and processor.processing_thread and processor.processing_thread.is_alive()
                has_recent_frame = (hasattr(processor, 'latest_raw_frame') and processor.latest_raw_frame is not None) or \
                                   (hasattr(processor, 'latest_annotated_frame') and processor.latest_annotated_frame is not None)
                is_actually_working = thread_alive or has_recent_frame
            
            if not is_actually_working:
                if no_frame_start_time is None:
                    no_frame_start_time = current_time
                elif current_time - no_frame_start_time > 3.0 and not error_emitted:
                    # Processor not working after 3 seconds - emit error
                    socketio.emit('stream_error', {
                        'app_name': app_name,
                        'channel_id': channel_id,
                        'error': f'Channel {channel_id} is not running. RTSP connection may have failed. Please check camera connectivity.'
                    }, room=None)
                    error_emitted = True
                time.sleep(0.5)
                continue
            
            # Reset error tracking if we have a processor and it's running
            if no_frame_start_time is not None:
                no_frame_start_time = None
                error_emitted = False
            
            # Get latest frame (either module-specific or combined)
            # Try module-specific first, then fall back to combined frame
            # Use timeout to avoid blocking too long
            frame = None
            try:
                active_modules = processor.get_active_modules()
                # Debug: Log active modules if empty or if MaterialTheftMonitor is requested
                if not active_modules or app_name == 'MaterialTheftMonitor':
                    logger.debug(f"📊 {app_name}/{channel_id}: Active modules = {active_modules}, processor.modules keys = {list(processor.modules.keys()) if hasattr(processor, 'modules') else 'N/A'}")
                
                if app_name in active_modules:
                    # Get module-specific frame
                    frame = processor.get_latest_frame(module_name=app_name)
                    # If module-specific frame is None or invalid, fall back to combined frame
                    if frame is None or (hasattr(frame, 'size') and frame.size == 0):
                        if consecutive_none_frames % 60 == 0:  # Log every 60th failure to avoid spam
                            logger.debug(f"Module-specific frame for {app_name}/{channel_id} not available, using combined frame")
                        # Fall back to combined frame
                        frame = processor.get_latest_frame(module_name=None)
                else:
                    # Module not in active modules - use combined frame as fallback
                    if consecutive_none_frames == 0:
                        logger.info(f"{app_name} not in active modules for {channel_id}. Active modules: {active_modules}. Using combined frame.")
                    # Use combined frame when module name doesn't match
                    frame = processor.get_latest_frame(module_name=None)
                
                # Final check - if still no frame, skip
                if frame is None or (hasattr(frame, 'size') and frame.size == 0):
                    if consecutive_none_frames % 60 == 0:  # Log every 60th failure to avoid spam
                        logger.debug(f"No frame available for {app_name}/{channel_id}, skipping")
                    time.sleep(0.05)
                    continue
            except Exception as e:
                logger.warning(f"Error getting frame for {stream_key}: {e}", exc_info=True)
                time.sleep(0.1)
                continue
            
            if frame is None:
                consecutive_none_frames += 1
                if consecutive_none_frames > max_consecutive_none:
                    if no_frame_start_time is None:
                        no_frame_start_time = current_time
                    elif current_time - no_frame_start_time > 3.0 and not error_emitted:
                        # No frames received after 3 seconds - emit error
                        socketio.emit('stream_error', {
                            'app_name': app_name,
                            'channel_id': channel_id,
                            'error': f'No video frames received from channel {channel_id}. Camera may be offline or RTSP stream unavailable.'
                        }, room=None)
                        error_emitted = True
                time.sleep(0.05)  # Shorter sleep for faster recovery
                continue
            
            # Reset counters on successful frame
            consecutive_none_frames = 0
            
            # Reset error tracking if we got a frame
            if no_frame_start_time is not None:
                no_frame_start_time = None
                error_emitted = False
            
            # Check frame freshness - skip if frame is too old (stale)
            # This prevents sending frozen/stuck frames
            frame_timestamp = getattr(frame, 'timestamp', None)
            if frame_timestamp and (current_time - frame_timestamp) > 2.0:
                logger.debug(f"Skipping stale frame for {stream_key} (age: {current_time - frame_timestamp:.2f}s)")
                time.sleep(0.05)
                continue
            
            # Resize frame if too large to reduce encoding time
            frame_height, frame_width = frame.shape[:2]
            if frame_width > 1280 or frame_height > 720:
                scale = min(1280 / frame_width, 720 / frame_height)
                new_width = int(frame_width * scale)
                new_height = int(frame_height * scale)
                frame = cv2.resize(frame, (new_width, new_height), interpolation=cv2.INTER_LINEAR)
            
            # Encode frame as JPEG (lower quality for faster encoding)
            ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
            if not ret:
                time.sleep(0.1)
                continue
            
            # Convert to base64 (optimized)
            try:
                frame_base64 = base64.b64encode(buffer).decode('utf-8')
            except Exception as e:
                logger.warning(f"Error encoding frame to base64 for {stream_key}: {e}")
                time.sleep(0.1)
                continue
            
            # Get FPS data
            fps_data = processor.get_live_fps() if hasattr(processor, 'get_live_fps') else {'live_feed_fps': 0, 'processing_fps': 0}
            
            # Broadcast to all subscribed clients
            socketio.emit('video_frame', {
                'app_name': app_name,
                'channel_id': channel_id,
                'frame': frame_base64,
                'timestamp': current_time,
                'fps': fps_data['live_feed_fps'],
                'processing_fps': fps_data['processing_fps']
            }, room=None)  # Broadcast to all connected clients
            
            last_frame_time = current_time
            
        except Exception as e:
            logger.error(f"Error broadcasting frame for {stream_key}: {e}")
            time.sleep(0.5)
    
    # Clean up
    if stream_key in stream_broadcast_threads:
        del stream_broadcast_threads[stream_key]
    if stream_key in stream_broadcast_stop_flags:
        del stream_broadcast_stop_flags[stream_key]
    
    logger.info(f"Stopped broadcasting frames for {stream_key}")

def create_database_tables():
    """Create database tables"""
    with app.app_context():
        db.create_all()
        logger.info("Database tables created")


# ============= Dress Code Monitoring Routes =============

@app.route('/static/dresscode_snapshots/<filename>')
def serve_dresscode_snapshot(filename):
    """Serve dress code violation snapshot"""
    return send_from_directory('static/dresscode_snapshots', filename)

@app.route('/api/get_dresscode_alerts')
def get_dresscode_alerts():
    """Get dress code violation alerts"""
    channel_id = request.args.get('channel_id')
    limit = int(request.args.get('limit', 50))
    
    try:
        with app.app_context():
            alerts = db_manager.get_dresscode_alerts(channel_id=channel_id, limit=limit)
        
        return jsonify({
            'success': True,
            'alerts': alerts,
            'count': len(alerts)
        })
    except Exception as e:
        logger.error(f"Error getting dress code alerts: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_counter_roi', methods=['POST'])
@login_required
def set_counter_roi():
    """Set counter ROI for dress code monitoring (uses best.pt for uniform detection)"""
    data = request.json
    channel_id = data.get('channel_id')
    roi_points = data.get('roi_points', [])
    
    if not channel_id:
        return jsonify({'success': False, 'error': 'channel_id is required'})
    
    try:
        # Get the DressCodeMonitoring module for this channel
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if 'DressCodeMonitoring' in processor.modules:
                module = processor.modules['DressCodeMonitoring']
                module.set_counter_roi(roi_points)
                return jsonify({'success': True, 'message': 'Counter ROI set successfully (best.pt for uniforms)'})
        
        return jsonify({'success': False, 'error': f'DressCodeMonitoring not found for channel {channel_id}'})
    except Exception as e:
        logger.error(f"Error setting counter ROI: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_queue_roi', methods=['POST'])
@login_required
def set_queue_roi():
    """Set queue ROI for dress code monitoring (uses YOLOv11 for person detection)"""
    data = request.json
    channel_id = data.get('channel_id')
    roi_points = data.get('roi_points', [])
    
    if not channel_id:
        return jsonify({'success': False, 'error': 'channel_id is required'})
    
    try:
        # Get the DressCodeMonitoring module for this channel
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if 'DressCodeMonitoring' in processor.modules:
                module = processor.modules['DressCodeMonitoring']
                module.set_queue_roi(roi_points)
                return jsonify({'success': True, 'message': 'Queue ROI set successfully (YOLOv11 for persons)'})
        
        return jsonify({'success': False, 'error': f'DressCodeMonitoring not found for channel {channel_id}'})
    except Exception as e:
        logger.error(f"Error setting queue ROI: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_table_roi', methods=['POST'])
@login_required
def set_table_roi():
    """Set table ROI for table service monitoring"""
    data = request.json
    channel_id = data.get('channel_id')
    table_id = data.get('table_id')
    roi_points = data.get('roi_points', [])
    
    if not channel_id or not table_id:
        return jsonify({'success': False, 'error': 'channel_id and table_id are required'})
    
    try:
        # Get the TableServiceMonitor module for this channel
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if 'TableServiceMonitor' in processor.modules:
                module = processor.modules['TableServiceMonitor']
                module.set_table_roi(table_id, roi_points)
                return jsonify({'success': True, 'message': f'Table ROI set successfully for table {table_id}'})
        
        return jsonify({'success': False, 'error': f'TableServiceMonitor not found for channel {channel_id}'})
    except Exception as e:
        logger.error(f"Error setting table ROI: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_service_discipline_table_roi', methods=['POST'])
@login_required
def set_service_discipline_table_roi():
    """Set table ROI for Service Discipline monitoring"""
    data = request.json
    channel_id = data.get('channel_id')
    table_id = data.get('table_id')
    roi_points = data.get('roi_points', [])
    
    if not channel_id or not table_id:
        return jsonify({'success': False, 'error': 'channel_id and table_id are required'})
    
    if not roi_points or len(roi_points) < 3:
        return jsonify({'success': False, 'error': 'At least 3 points are required for a polygon'})
    
    try:
        module = None
        
        # Try to get module from channel_modules first
        if channel_id in channel_modules and 'ServiceDisciplineMonitor' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['ServiceDisciplineMonitor']
        # Also try from shared processor
        elif channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if hasattr(processor, 'modules') and 'ServiceDisciplineMonitor' in processor.modules:
                module = processor.modules['ServiceDisciplineMonitor']
            elif hasattr(processor, 'get_module'):
                module = processor.get_module('ServiceDisciplineMonitor')
        
        if module and hasattr(module, 'set_table_roi'):
            # Normalize roi_points format - ensure they're in the right format
            normalized_points = []
            for point in roi_points:
                if isinstance(point, dict):
                    normalized_points.append(point)
                elif isinstance(point, (list, tuple)) and len(point) >= 2:
                    normalized_points.append({'x': float(point[0]), 'y': float(point[1])})
                else:
                    return jsonify({'success': False, 'error': f'Invalid point format: {point}'})
            
            module.set_table_roi(table_id, normalized_points)
            
            # Reload configuration to ensure it's persisted and loaded correctly
            if hasattr(module, 'load_configuration'):
                try:
                    module.load_configuration()
                    logger.info(f"✓ Configuration reloaded after setting ROI for {table_id}")
                except Exception as e:
                    logger.warning(f"Could not reload configuration: {e}")
            
            logger.info(f"✓ Table ROI set for {table_id} on channel {channel_id} (total tables: {len(module.table_rois)})")
            return jsonify({'success': True, 'message': f'Table ROI set successfully for table {table_id}'})
        
        # If module not found, provide detailed error
        available_modules = []
        if channel_id in channel_modules:
            available_modules = list(channel_modules[channel_id].keys())
        elif channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if hasattr(processor, 'modules'):
                available_modules = list(processor.modules.keys())
        
        error_msg = f'ServiceDisciplineMonitor not found for channel {channel_id}'
        if available_modules:
            error_msg += f'. Available modules: {", ".join(available_modules)}'
        else:
            error_msg += f'. Channel {channel_id} not found in active channels.'
        
        logger.error(error_msg)
        return jsonify({'success': False, 'error': error_msg})
    except Exception as e:
        logger.error(f"Error setting service discipline table ROI: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_service_discipline_table_rois', methods=['GET'])
@login_required
def get_service_discipline_table_rois():
    """Get all table ROIs for Service Discipline monitoring"""
    channel_id = request.args.get('channel_id')
    
    if not channel_id:
        return jsonify({'success': False, 'error': 'channel_id is required'})
    
    try:
        module = None
        
        # Try to get module from channel_modules first
        if channel_id in channel_modules and 'ServiceDisciplineMonitor' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['ServiceDisciplineMonitor']
        # Also try from shared processor
        elif channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if hasattr(processor, 'modules') and 'ServiceDisciplineMonitor' in processor.modules:
                module = processor.modules['ServiceDisciplineMonitor']
            elif hasattr(processor, 'get_module'):
                module = processor.get_module('ServiceDisciplineMonitor')
        
        if module and hasattr(module, 'table_rois'):
            table_rois = module.table_rois.copy()
            # Convert to serializable format
            result = {}
            for table_id, roi_info in table_rois.items():
                result[table_id] = {
                    'polygon': roi_info.get('polygon', []),
                    'bbox': roi_info.get('bbox', (0, 0, 0, 0))
                }
            return jsonify({'success': True, 'table_rois': result})
        
        return jsonify({'success': True, 'table_rois': {}})
    except Exception as e:
        logger.error(f"Error getting service discipline table ROIs: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_table_service_violations')
@login_required
def get_table_service_violations():
    """Get table service violations"""
    channel_id = request.args.get('channel_id')
    violation_type = request.args.get('violation_type')  # 'service_discipline' or 'unclean_table' or 'slow_reset'
    limit = int(request.args.get('limit', 50))
    days = request.args.get('days')
    days = int(days) if days else None
    
    try:
        with app.app_context():
            violations = db_manager.get_table_service_violations(
                channel_id=channel_id, 
                violation_type=violation_type,
                limit=limit,
                days=days
            )
        
        return jsonify({
            'success': True,
            'violations': violations,
            'count': len(violations)
        })
    except Exception as e:
        logger.error(f"Error getting table service violations: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_table_cleanliness_violations')
@login_required
def get_table_cleanliness_violations():
    """Get table cleanliness violations (unclean/slow reset only)"""
    channel_id = request.args.get('channel_id')
    table_id = request.args.get('table_id')
    limit = int(request.args.get('limit', 50))
    days = request.args.get('days')
    days = int(days) if days else None

    try:
        with app.app_context():
            violations = db_manager.get_table_cleanliness_violations(
                channel_id=channel_id,
                table_id=table_id,
                limit=limit,
                days=days,
            )
        return jsonify({'success': True, 'violations': violations, 'count': len(violations)})
    except Exception as e:
        logger.error(f"Error getting table cleanliness violations: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_table_cleanliness_violation', methods=['POST'])
@login_required
def delete_table_cleanliness_violation():
    """Delete a table cleanliness violation"""
    data = request.json
    violation_id = data.get('violation_id')

    if not violation_id:
        return jsonify({'success': False, 'error': 'violation_id is required'})

    try:
        with app.app_context():
            success = db_manager.delete_table_cleanliness_violation(violation_id)
            if success:
                return jsonify({'success': True, 'message': 'Violation deleted successfully'})
            return jsonify({'success': False, 'error': 'Violation not found'})
    except Exception as e:
        logger.error(f"Error deleting table cleanliness violation: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_table_service_violation', methods=['POST'])
@login_required
def delete_table_service_violation():
    """Delete a table service violation"""
    data = request.json
    violation_id = data.get('violation_id')
    
    if not violation_id:
        return jsonify({'success': False, 'error': 'violation_id is required'})
    
    try:
        with app.app_context():
            success = db_manager.delete_table_service_violation(violation_id)
            if success:
                return jsonify({'success': True, 'message': 'Violation deleted successfully'})
            else:
                return jsonify({'success': False, 'error': 'Violation not found'})
    except Exception as e:
        logger.error(f"Error deleting table service violation: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_all_table_service_violations', methods=['POST'])
@login_required
def clear_all_table_service_violations():
    """Delete ALL table service violations (use with caution!)"""
    try:
        with app.app_context():
            deleted_count = db_manager.clear_all_table_service_violations()
            return jsonify({
                'success': True,
                'deleted_count': deleted_count,
                'message': f'Deleted {deleted_count} service discipline violations'
            })
    except Exception as e:
        logger.error(f"Error clearing all table service violations: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_old_table_service_violations', methods=['POST'])
@login_required
def clear_old_table_service_violations():
    """Clear old table service violations"""
    data = request.json
    days = data.get('days', 7)
    
    try:
        with app.app_context():
            deleted_count = db_manager.clear_old_table_service_violations(days=days)
            return jsonify({
                'success': True,
                'deleted_count': deleted_count,
                'message': f'Deleted {deleted_count} violations older than {days} days'
            })
    except Exception as e:
        logger.error(f"Error clearing old table service violations: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/set_allowed_uniforms', methods=['POST'])
@login_required
def set_allowed_uniforms():
    """Set allowed uniform colors for counter area"""
    data = request.json
    channel_id = data.get('channel_id')
    allowed_uniforms = data.get('allowed_uniforms', {})
    
    if not channel_id:
        return jsonify({'success': False, 'error': 'channel_id is required'})
    
    try:
        # Get the DressCodeMonitoring module for this channel
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if 'DressCodeMonitoring' in processor.modules:
                module = processor.modules['DressCodeMonitoring']
                module.set_allowed_uniforms(allowed_uniforms)
                return jsonify({'success': True, 'message': 'Allowed uniforms set successfully'})
        
        return jsonify({'success': False, 'error': f'DressCodeMonitoring not found for channel {channel_id}'})
    except Exception as e:
        logger.error(f"Error setting allowed uniforms: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_counter_roi/<channel_id>')
@login_required
def get_counter_roi(channel_id):
    """Get counter and queue ROI configuration"""
    try:
        if channel_id in shared_video_processors:
            processor = shared_video_processors[channel_id]
            if 'DressCodeMonitoring' in processor.modules:
                module = processor.modules['DressCodeMonitoring']
                return jsonify({
                    'success': True,
                    'counter_roi': module.counter_roi,
                    'queue_roi': module.queue_roi,
                    'allowed_uniforms': module.allowed_uniforms_in_counter
                })
        
        return jsonify({'success': False, 'error': f'DressCodeMonitoring not found for channel {channel_id}'})
    except Exception as e:
        logger.error(f"Error getting ROI configuration: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_dresscode_alert/<int:alert_id>', methods=['DELETE'])
@login_required
def delete_dresscode_alert(alert_id):
    """Delete a dress code alert"""
    try:
        with app.app_context():
            success = db_manager.delete_dresscode_alert(alert_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Alert deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Alert not found'})
    except Exception as e:
        logger.error(f"Error deleting dress code alert: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_dresscode_stats')
def get_dresscode_stats():
    """Get dress code violation statistics"""
    channel_id = request.args.get('channel_id')
    days = int(request.args.get('days', 7))
    
    try:
        with app.app_context():
            stats = db_manager.get_dresscode_stats(channel_id=channel_id, days=days)
        
        return jsonify({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        logger.error(f"Error getting dress code stats: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_ppe_alerts')
def get_ppe_alerts():
    """Get PPE violation alerts"""
    channel_id = request.args.get('channel_id')
    limit = int(request.args.get('limit', 50))
    
    try:
        with app.app_context():
            alerts = db_manager.get_ppe_alerts(channel_id=channel_id, limit=limit)
        
        return jsonify({
            'success': True,
            'alerts': alerts,
            'count': len(alerts)
        })
    except Exception as e:
        logger.error(f"Error getting PPE alerts: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_ppe_alert/<int:alert_id>', methods=['DELETE'])
@login_required
def delete_ppe_alert(alert_id):
    """Delete a PPE violation alert"""
    try:
        with app.app_context():
            success = db_manager.delete_ppe_alert(alert_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Alert deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Alert not found'})
    except Exception as e:
        logger.error(f"Error deleting PPE alert: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_ppe_stats')
def get_ppe_stats():
    """Get PPE violation statistics"""
    channel_id = request.args.get('channel_id')
    days = int(request.args.get('days', 7))
    
    try:
        with app.app_context():
            stats = db_manager.get_ppe_stats(channel_id=channel_id, days=days)
        
        return jsonify({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        logger.error(f"Error getting PPE stats: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_queue_violations')
def get_queue_violations():
    """Get queue violation alerts"""
    channel_id = request.args.get('channel_id')
    limit = int(request.args.get('limit', 50))
    
    try:
        with app.app_context():
            violations = db_manager.get_queue_violations(channel_id=channel_id, limit=limit)
        
        logger.info(f"Retrieved {len(violations)} queue violations (channel_id={channel_id}, limit={limit})")
        
        # Ensure each violation has a timestamp field (for compatibility)
        for violation in violations:
            if 'timestamp' not in violation and 'created_at' in violation:
                violation['timestamp'] = violation['created_at']
        
        return jsonify({
            'success': True,
            'violations': violations,
            'count': len(violations)
        })
    except Exception as e:
        logger.error(f"Error getting queue violations: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_queue_violation/<int:violation_id>', methods=['DELETE'])
@login_required
def delete_queue_violation(violation_id):
    """Delete a queue violation"""
    try:
        with app.app_context():
            success = db_manager.delete_queue_violation(violation_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Violation deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Violation not found'})
    except Exception as e:
        logger.error(f"Error deleting queue violation: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/update_dresscode_config', methods=['POST'])
@login_required
def update_dresscode_config():
    """Update dress code monitoring configuration"""
    data = request.json
    channel_id = data.get('channel_id')
    config = data.get('config', {})
    
    try:
        # Get the module
        if channel_id in channel_modules and 'DressCodeMonitoring' in channel_modules[channel_id]:
            module = channel_modules[channel_id]['DressCodeMonitoring']
            
            # Update configuration
            if 'alert_cooldown' in config:
                module.alert_cooldown = float(config['alert_cooldown'])
            if 'violation_duration_threshold' in config:
                module.violation_duration_threshold = float(config['violation_duration_threshold'])
            if 'conf_threshold' in config:
                module.conf_threshold = float(config['conf_threshold'])
            
            return jsonify({
                'success': True,
                'message': 'Configuration updated successfully'
            })
        else:
            return jsonify({'success': False, 'error': 'Module not found'})
            
    except Exception as e:
        logger.error(f"Error updating dress code config: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/system_resources')
@login_required
def get_system_resources():
    """Get current system resource usage"""
    try:
        import psutil
        resources = {
            'cpu_percent': psutil.cpu_percent(),
            'memory_percent': psutil.virtual_memory().percent,
            'available_memory_gb': psutil.virtual_memory().available / (1024**3)
        }
        return jsonify({
            'success': True,
            'resources': resources,
            'active_channels': len(shared_video_processors),
            'max_channels': 50
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

if __name__ == '__main__':
    # Create database tables
    create_database_tables()
    
    # Start the application FIRST (non-blocking)
    logger.info("Starting Sakshi.AI Video Analytics Platform")
    logger.info("Dashboard will be available at http://localhost:5000")
    
    # Load and start channels from configuration in BACKGROUND (non-blocking)
    def load_channels_in_background():
        """Load channels in background thread so server starts immediately"""
        time.sleep(1)  # Give server a moment to start
        logger.info("Loading channels from configuration in background...")
        load_channels_from_config()
        logger.info("✅ Channel loading completed")
    
    # Start channel loading in background thread
    channel_loader_thread = threading.Thread(target=load_channels_in_background, daemon=True)
    channel_loader_thread.start()
    
    # Start Flask server (this will block, but server is now running)
    socketio.run(app, host='0.0.0.0', port=5000, debug=False, allow_unsafe_werkzeug=True)